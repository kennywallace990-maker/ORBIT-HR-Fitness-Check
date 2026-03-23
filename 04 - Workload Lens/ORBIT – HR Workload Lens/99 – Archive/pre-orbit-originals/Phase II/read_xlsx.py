import openpyxl
import os

path = os.path.join(os.path.dirname(__file__), "Phase II CSVs", "WoW_Variance_Analysis_Week10 v1 ticket data only.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
print("Sheets:", wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n=== Sheet: {sname} (rows={ws.max_row}, cols={ws.max_column}) ===")
    for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=False):
        vals = []
        for c in row:
            comment_text = c.comment.text.strip() if c.comment else None
            if comment_text:
                vals.append(f"{c.value} [COMMENT: {comment_text}]")
            else:
                vals.append(c.value)
        print("\t".join(str(v) for v in vals))
