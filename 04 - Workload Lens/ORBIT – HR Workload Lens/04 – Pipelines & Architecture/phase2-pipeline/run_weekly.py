"""
Workload Lens — Weekly Pipeline Runner
=======================================
Usage:
  1. Drop this week's ServiceNow CSVs into  Phase II CSVs/
     Expected files (naming flexible — script auto-detects by Hr Service column):
       - Attendance Inquiry *.csv
       - CC Time and Attendance *.csv   (or "CS Time and Attendance")
       - FC General Inquiry *.csv
       - Timesheet Inquiry *.csv
  2. Drop the OBR PDF into  Phase II CSVs/  (optional — used for validation)
  3. Run:  python run_weekly.py
  4. Report appears in  output/

The script will:
  - Auto-detect CSVs and the two most recent weeks of data
  - Classify every ticket using classifier.py rules
  - Extract OBR text from PDF (if present)
  - Generate the Insights & Recommendations HTML report
"""
import csv
import os
import re
import sys
import json
from datetime import datetime, timedelta
from collections import Counter, defaultdict

# Add pipeline dir to path so we can import classifier
PIPELINE_DIR = os.path.dirname(os.path.abspath(__file__))
PHASE2_DIR = os.path.dirname(PIPELINE_DIR)
sys.path.insert(0, PIPELINE_DIR)

from classifier import classify_ticket, get_ss_channel

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
CSV_DIR = os.path.join(PHASE2_DIR, "Phase II CSVs")
OUTPUT_DIR = os.path.join(PHASE2_DIR, "output")
REPORT_TEMPLATE = os.path.join(PIPELINE_DIR, "report_template.html")
SELF_SERVICE_POTENTIAL_ASSUMPTION = 0.80
CUSTOM_REPORT_WEEK = "2026-03-01"
CUSTOM_REPORT_LABEL = "Week 10"
CUSTOM_REPORT_TOTAL_WEEK10_VOLUME = 7514
CUSTOM_REPORT_UKG_ACTIONS_PER_WEEK = 67543
CUSTOM_REPORT_UKG_REWORK_ACTIONS_PER_WEEK = 22885
CUSTOM_REPORT_UKG_FTE_HOURS_PER_WEEK = 761

# Services we track (normalized names)
SERVICES = [
    "Attendance Inquiry",
    "CC Time and Attendance",
    "FC General Inquiry",
    "Timesheet Inquiry",
]

# Service name normalization map
SERVICE_NORMALIZE = {
    "attendance inquiry": "Attendance Inquiry",
    "cs time and attendance": "CC Time and Attendance",
    "cc time and attendance": "CC Time and Attendance",
    "fc general inquiry": "FC General Inquiry",
    "timesheet inquiry": "Timesheet Inquiry",
}


def parse_date(s):
    if not s:
        return None
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # Try datetime with time
    for fmt in ("%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def get_week_sunday(dt):
    """Return the Sunday that starts the week containing dt."""
    return dt - timedelta(days=dt.weekday() + 1) if dt.weekday() != 6 else dt


def extract_site(assignment_group):
    if not assignment_group:
        return "UNKNOWN"
    ag = assignment_group.strip()
    centralized = [
        "Real Time Analyst Tier I", "Real Time Analyst Tier II",
        "HR Team Member Data Management", "HR Team Member Service Center",
        "LOA/ADA Team", "Payroll Team", "HR TMDM",
    ]
    for c in centralized:
        if c.lower() in ag.lower():
            return ag.strip()
    m = re.match(r'^([A-Z]{2,4}\s?\d{0,2}(?:/\d+)*(?:\s?HR)?)', ag)
    if m:
        return m.group(1).strip()
    return ag[:30]


def extract_obr_text(csv_dir):
    """Find and extract text from OBR PDF if present."""
    pdf_files = [f for f in os.listdir(csv_dir) if f.lower().endswith('.pdf')]
    if not pdf_files:
        return None
    pdf_path = os.path.join(csv_dir, pdf_files[0])
    try:
        import pdfplumber
        text_pages = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text_pages.append(page.extract_text() or "")
        return "\n".join(text_pages)
    except Exception as e:
        print(f"  Warning: Could not extract PDF text: {e}")
        return None


def load_csvs(csv_dir):
    """Load all CSVs, deduplicate, return list of ticket dicts."""
    csv_files = [f for f in os.listdir(csv_dir) if f.lower().endswith('.csv')]
    if not csv_files:
        print(f"ERROR: No CSV files found in {csv_dir}")
        sys.exit(1)

    all_tickets = []
    seen_numbers = set()
    files_loaded = []

    for fname in csv_files:
        fpath = os.path.join(csv_dir, fname)
        try:
            with open(fpath, 'r', encoding='utf-8-sig', errors='replace') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        except Exception as e:
            print(f"  Warning: Could not read {fname}: {e}")
            continue

        count = 0
        for row in rows:
            num = (row.get('Number') or '').strip()
            if not num or not num.startswith('HRC'):
                continue
            if num in seen_numbers:
                continue
            seen_numbers.add(num)

            hr_service_raw = (row.get('Hr Service') or row.get('\ufeffHr Service') or '').strip()
            hr_service = SERVICE_NORMALIZE.get(hr_service_raw.lower(), hr_service_raw)
            if hr_service not in SERVICES:
                continue

            opened = parse_date(row.get('Opened At', ''))
            resolved = parse_date(row.get('U Resolved', ''))
            desc = (row.get('Description1') or '')[:2000]
            ag = (row.get('Assignment Group') or '').strip()
            contact = (row.get('Contact Type') or '').strip()
            site = extract_site(ag)

            all_tickets.append({
                'number': num,
                'service': hr_service,
                'opened': opened,
                'resolved': resolved,
                'description': desc,
                'assignment_group': ag,
                'contact_type': contact,
                'site': site,
            })
            count += 1

        if count > 0:
            files_loaded.append(f"{fname} ({count} tickets)")

    return all_tickets, files_loaded


def detect_weeks(tickets):
    """Auto-detect the two most recent full weeks (Sun-Sat) in the data."""
    dates = [t['opened'] for t in tickets if t['opened']]
    if not dates:
        print("ERROR: No valid dates found in ticket data.")
        sys.exit(1)

    week_starts = Counter()
    for d in dates:
        ws = get_week_sunday(d)
        ws_date = ws.replace(hour=0, minute=0, second=0, microsecond=0)
        week_starts[ws_date] += 1

    # Get the two most common week starts (most recent)
    top_weeks = sorted(week_starts.keys(), reverse=True)
    if len(top_weeks) < 2:
        print("ERROR: Need at least 2 weeks of data. Found:", len(top_weeks))
        sys.exit(1)

    # Take the two most recent with significant volume
    current_week = top_weeks[0]
    prior_week = top_weeks[1]

    return prior_week, current_week


def assign_weeks(tickets, prior_start, current_start):
    """Assign each ticket to Prior Week or Current Week."""
    prior_end = prior_start + timedelta(days=6)
    current_end = current_start + timedelta(days=6)

    for t in tickets:
        if not t['opened']:
            t['week'] = None
            continue
        d = t['opened'].replace(hour=0, minute=0, second=0, microsecond=0)
        if prior_start <= d <= prior_end:
            t['week'] = 'prior'
        elif current_start <= d <= current_end:
            t['week'] = 'current'
        else:
            t['week'] = None


def classify_all(tickets):
    """Run classifier on all tickets."""
    for t in tickets:
        classification, sub_cat = classify_ticket(
            t['description'], t['contact_type'], t['service'],
            t.get('assignment_group', '')
        )
        t['classification'] = classification
        t['sub_category'] = sub_cat
        t['ss_channel'] = get_ss_channel(sub_cat)


def compute_metrics(tickets, prior_start, current_start):
    """Compute all metrics needed for the report."""
    prior_label = prior_start.strftime('%b %d')
    prior_end_label = (prior_start + timedelta(days=6)).strftime('%b %d')
    current_label = current_start.strftime('%b %d')
    current_end_label = (current_start + timedelta(days=6)).strftime('%b %d')

    # Filter to only assigned weeks
    prior = [t for t in tickets if t.get('week') == 'prior']
    current = [t for t in tickets if t.get('week') == 'current']
    both = prior + current

    total_prior = len(prior)
    total_current = len(current)
    total_all = len(both)

    # ── Classification breakdown ──
    class_counts = defaultdict(lambda: {'total': 0, 'prior': 0, 'current': 0})
    for t in both:
        c = t['classification']
        class_counts[c]['total'] += 1
        class_counts[c][t['week']] += 1

    # ── Service breakdown ──
    svc_counts = defaultdict(lambda: {'prior': 0, 'current': 0})
    for t in both:
        svc_counts[t['service']][t['week']] += 1

    # ── Self-service sub-category breakdown ──
    ss_subcats = defaultdict(lambda: {'prior': 0, 'current': 0, 'channel': ''})
    for t in both:
        if t['classification'] == 'Self Service Eligible':
            sc = t['sub_category']
            ss_subcats[sc][t['week']] += 1
            ss_subcats[sc]['channel'] = t['ss_channel']

    # ── Process Required sub-category breakdown ──
    pr_subcats = defaultdict(lambda: {'prior': 0, 'current': 0})
    for t in both:
        if t['classification'] == 'Process Required':
            pr_subcats[t['sub_category']][t['week']] += 1

    # ── Defect sub-category breakdown ──
    def_subcats = defaultdict(lambda: {'prior': 0, 'current': 0})
    for t in both:
        if t['classification'] == 'Defect':
            def_subcats[t['sub_category']][t['week']] += 1

    # ── Process Optimization sub-category breakdown ──
    po_subcats = defaultdict(lambda: {'prior': 0, 'current': 0})
    for t in both:
        if t['classification'] == 'Process Optimization':
            po_subcats[t['sub_category']][t['week']] += 1

    # ── Site/Assignment Group breakdown ──
    site_counts = defaultdict(lambda: {
        'prior': 0, 'current': 0, 'ss_prior': 0, 'ss_current': 0
    })
    for t in both:
        ag = t['assignment_group'] or t['site']
        site_counts[ag][t['week']] += 1
        if t['classification'] == 'Self Service Eligible':
            site_counts[ag][f'ss_{t["week"]}'] += 1

    # ── Service x Classification breakdown ──
    svc_class = defaultdict(lambda: defaultdict(lambda: {'prior': 0, 'current': 0}))
    for t in both:
        svc_class[t['service']][t['classification']][t['week']] += 1

    # ── Contact type breakdown (for Attendance Inquiry) ──
    attend_contact = defaultdict(lambda: {'prior': 0, 'current': 0})
    for t in both:
        if t['service'] == 'Attendance Inquiry':
            ct = t['contact_type'] or 'Unknown'
            attend_contact[ct][t['week']] += 1

    return {
        'prior_label': f"{prior_label} – {prior_end_label}",
        'current_label': f"{current_label} – {current_end_label}",
        'prior_start': prior_start,
        'current_start': current_start,
        'total_prior': total_prior,
        'total_current': total_current,
        'total_all': total_all,
        'class_counts': dict(class_counts),
        'svc_counts': dict(svc_counts),
        'ss_subcats': dict(ss_subcats),
        'pr_subcats': dict(pr_subcats),
        'def_subcats': dict(def_subcats),
        'po_subcats': dict(po_subcats),
        'site_counts': dict(site_counts),
        'svc_class': {k: dict(v) for k, v in svc_class.items()},
        'attend_contact': dict(attend_contact),
    }


def pct(n, d):
    return f"{n/d*100:.1f}%" if d else "0%"


def delta_str(prior, current):
    d = current - prior
    p = (d / prior * 100) if prior else 0
    sign = "+" if d >= 0 else ""
    color = "down" if d < 0 else ("up" if d > 0 else "")
    return d, p, f'{sign}{d} ({sign}{p:.1f}%)', color


def generate_custom_week_report(metrics, output_path):
    """Temporary custom report for the Week 10 leadership readout."""
    m = metrics
    prior_n = m['total_prior']
    curr_n = m['total_current']
    total_n = m['total_all']
    wow_delta, wow_pct, wow_str, wow_color = delta_str(prior_n, curr_n)
    wow_abs = abs(wow_delta)
    prior_lbl = m['prior_label']
    curr_lbl = m['current_label']
    report_date = datetime.now().strftime('%B %d, %Y')
    attendance_current = m['svc_counts'].get('Attendance Inquiry', {}).get('current', 0)
    attendance_prior = m['svc_counts'].get('Attendance Inquiry', {}).get('prior', 0)
    attendance_delta = attendance_current - attendance_prior
    attendance_validated_prior = (
        m['svc_class'].get('Attendance Inquiry', {}).get('Process Required', {}).get('prior', 0)
        + m['svc_class'].get('Attendance Inquiry', {}).get('Defect', {}).get('prior', 0)
    )
    attendance_validated_current = (
        m['svc_class'].get('Attendance Inquiry', {}).get('Process Required', {}).get('current', 0)
        + m['svc_class'].get('Attendance Inquiry', {}).get('Defect', {}).get('current', 0)
    )
    attendance_validated_drop = max(0, attendance_validated_prior - attendance_validated_current)
    fc_current = m['svc_counts'].get('FC General Inquiry', {}).get('current', 0)
    fc_prior = m['svc_counts'].get('FC General Inquiry', {}).get('prior', 0)
    fc_delta = fc_current - fc_prior
    cc_current = m['svc_counts'].get('CC Time and Attendance', {}).get('current', 0)

    current_service_sorted = sorted(
        SERVICES,
        key=lambda svc: m['svc_counts'].get(svc, {}).get('current', 0),
        reverse=True,
    )
    largest_service = current_service_sorted[0] if current_service_sorted else SERVICES[0]
    largest_service_current = m['svc_counts'].get(largest_service, {}).get('current', 0)
    largest_service_pct = largest_service_current / curr_n * 100 if curr_n else 0

    validated_classes = ['Process Optimization', 'Process Required', 'Defect', 'Unclear']
    class_rows = ""
    for name in validated_classes:
        data = m['class_counts'].get(name, {'prior': 0, 'current': 0})
        prior = data.get('prior', 0)
        current = data.get('current', 0)
        _, _, delta_text, delta_class = delta_str(prior, current)
        pct_current = f"{current/curr_n*100:.1f}%" if curr_n else "0%"
        class_rows += (
            f'<tr><td>{name}</td><td class="r">{current:,}</td><td class="r">{pct_current}</td>'
            f'<td class="r">{prior:,}</td><td class="r {delta_class}">{delta_text}</td></tr>'
        )
    class_rows += (
        '<tr><td>Self-service candidate model</td>'
        '<td colspan="4">Training and validation are in progress. Metrics are intentionally withheld this week and will be published starting next week.</td></tr>'
    )
    class_rows += (
        f'<tr class="total"><td>Total Week 10 Volume</td><td class="r">{curr_n:,}</td><td class="r">100.0%</td>'
        f'<td class="r">{prior_n:,}</td><td class="r {wow_color}">{wow_str}</td></tr>'
    )

    service_rows = ""
    for svc in SERVICES:
        prior = m['svc_counts'].get(svc, {}).get('prior', 0)
        current = m['svc_counts'].get(svc, {}).get('current', 0)
        _, _, delta_text, delta_class = delta_str(prior, current)
        pct_current = f"{current/curr_n*100:.1f}%" if curr_n else "0%"
        service_rows += (
            f'<tr><td>{svc}</td><td class="r">{current:,}</td><td class="r">{pct_current}</td>'
            f'<td class="r">{prior:,}</td><td class="r {delta_class}">{delta_text}</td></tr>'
        )
    service_rows += (
        f'<tr class="total"><td>Total</td><td class="r">{curr_n:,}</td><td class="r">100.0%</td>'
        f'<td class="r">{prior_n:,}</td><td class="r {wow_color}">{wow_str}</td></tr>'
    )

    pr_current_total = m['class_counts'].get('Process Required', {}).get('current', 0)
    pr_sorted_current = sorted(
        m['pr_subcats'].items(),
        key=lambda item: item[1].get('current', 0),
        reverse=True,
    )
    pr_rows = ""
    pr_shown = 0
    pr_shown_current = 0
    pr_shown_prior = 0
    for i, (subcat, data) in enumerate(pr_sorted_current):
        prior = data.get('prior', 0)
        current = data.get('current', 0)
        if i < 8:
            _, _, delta_text, delta_class = delta_str(prior, current)
            pct_current = f"{current/pr_current_total*100:.1f}%" if pr_current_total else "0%"
            pr_rows += (
                f'<tr><td>{subcat}</td><td class="r">{current:,}</td><td class="r">{pct_current}</td>'
                f'<td class="r">{prior:,}</td><td class="r {delta_class}">{delta_text}</td></tr>'
            )
            pr_shown += 1
            pr_shown_current += current
            pr_shown_prior += prior
    other_pr_current = max(0, pr_current_total - pr_shown_current)
    other_pr_prior = max(0, m['class_counts'].get('Process Required', {}).get('prior', 0) - pr_shown_prior)
    if other_pr_current or other_pr_prior:
        _, _, delta_text, delta_class = delta_str(other_pr_prior, other_pr_current)
        pct_current = f"{other_pr_current/pr_current_total*100:.1f}%" if pr_current_total else "0%"
        pr_rows += (
            f'<tr><td>All other process-required</td><td class="r">{other_pr_current:,}</td><td class="r">{pct_current}</td>'
            f'<td class="r">{other_pr_prior:,}</td><td class="r {delta_class}">{delta_text}</td></tr>'
        )
    pr_rows += (
        f'<tr class="total"><td>Total Process Required</td><td class="r">{pr_current_total:,}</td><td class="r">100.0%</td>'
        f'<td class="r">{m["class_counts"].get("Process Required", {}).get("prior", 0):,}</td>'
        f'<td class="r"></td></tr>'
    )

    def_current_total = m['class_counts'].get('Defect', {}).get('current', 0)
    def_sorted_current = sorted(
        m['def_subcats'].items(),
        key=lambda item: item[1].get('current', 0),
        reverse=True,
    )
    top_def_name, top_def_data = def_sorted_current[0] if def_sorted_current else ("No dominant defect identified", {'prior': 0, 'current': 0})
    top_def_current = top_def_data.get('current', 0)
    top_def_pct_current = f"{top_def_current/def_current_total*100:.1f}%" if def_current_total else "0%"

    po_prior = m['class_counts'].get('Process Optimization', {}).get('prior', 0)
    po_current = m['class_counts'].get('Process Optimization', {}).get('current', 0)
    po_pct_current = po_current / curr_n * 100 if curr_n else 0

    site_sorted_current = sorted(
        m['site_counts'].items(),
        key=lambda item: item[1].get('current', 0),
        reverse=True,
    )
    top_sites_rows = ""
    top_sites_total = 0
    for assignment_group, data in site_sorted_current[:5]:
        prior = data.get('prior', 0)
        current = data.get('current', 0)
        top_sites_total += current
        _, _, delta_text, delta_class = delta_str(prior, current)
        pct_current = f"{current/curr_n*100:.1f}%" if curr_n else "0%"
        top_sites_rows += (
            f'<tr><td>{assignment_group}</td><td class="r">{current:,}</td><td class="r">{pct_current}</td>'
            f'<td class="r">{prior:,}</td><td class="r {delta_class}">{delta_text}</td></tr>'
        )
    top_sites_pct = f"{top_sites_total/curr_n*100:.1f}%" if curr_n else "0%"
    top_sites_rows += (
        f'<tr class="total"><td>Top 5 Subtotal</td><td class="r">{top_sites_total:,}</td><td class="r">{top_sites_pct}</td>'
        f'<td class="r"></td><td class="r"></td></tr>'
    )

    smartsheet_current = m['pr_subcats'].get('Smartsheet automated notification', {}).get('current', 0)
    smartsheet_pct = f"{smartsheet_current/pr_current_total*100:.1f}%" if pr_current_total else "0%"
    enterprise_slice_pct = curr_n / CUSTOM_REPORT_TOTAL_WEEK10_VOLUME * 100 if CUSTOM_REPORT_TOTAL_WEEK10_VOLUME else 0
    visible_vs_all_work_pct = curr_n / CUSTOM_REPORT_UKG_ACTIONS_PER_WEEK * 100 if CUSTOM_REPORT_UKG_ACTIONS_PER_WEEK else 0
    visible_vs_rework_pct = curr_n / CUSTOM_REPORT_UKG_REWORK_ACTIONS_PER_WEEK * 100 if CUSTOM_REPORT_UKG_REWORK_ACTIONS_PER_WEEK else 0
    issue_deltas = []
    for classification_name, subcats in (
        ('Process Optimization', m['po_subcats']),
        ('Process Required', m['pr_subcats']),
        ('Defect', m['def_subcats']),
    ):
        for name, data in subcats.items():
            delta = data.get('current', 0) - data.get('prior', 0)
            if delta:
                issue_deltas.append({
                    'classification': classification_name,
                    'name': name,
                    'delta': delta,
                })

    negative_issue_deltas = sorted(
        (item for item in issue_deltas if item['delta'] < 0),
        key=lambda item: abs(item['delta']),
        reverse=True,
    )
    positive_issue_deltas = sorted(
        (item for item in issue_deltas if item['delta'] > 0),
        key=lambda item: abs(item['delta']),
        reverse=True,
    )

    def format_issue_delta(item):
        return f"{item['name']} ({item['delta']:+,})"

    negative_issue_summary = ", ".join(format_issue_delta(item) for item in negative_issue_deltas[:4])
    positive_issue_summary = ", ".join(format_issue_delta(item) for item in positive_issue_deltas[:3])
    negative_issue_summary_topline = ", ".join(
        format_issue_delta(item)
        for item in [
            issue for issue in negative_issue_deltas
            if 'fmla' not in issue['name'].lower() and 'loa' not in issue['name'].lower()
        ][:4]
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>HR Workload Lens &mdash; {CUSTOM_REPORT_LABEL} Insights</title>
  <style>
    :root {{--ink:#1a1f2b;--body:#3a3f4c;--muted:#717888;--border:#e2e6ee;--panel:#ffffff;--bg:#f7f8fb;--blue:#0d5f73;--blue-soft:#e4f3f7;--blue-mid:#b0dce8;--green:#14643a;--green-bg:#e6f4ea;--red:#9b2c2c;--red-bg:#fdf0f0;--warn:#7a3e00;--warn-bg:#fff7ed;--purple:#5b21b6;--purple-bg:#f3edff;--shadow:0 4px 20px rgba(20,30,50,0.06);}}
    *{{box-sizing:border-box;margin:0;}}
    body{{font-family:Aptos,"Segoe UI",Calibri,sans-serif;color:var(--body);background:var(--bg);line-height:1.68;font-size:15.5px;}}
    .wrap{{max-width:860px;margin:0 auto;padding:40px 28px 64px;}}
    .header{{background:var(--ink);color:#fff;padding:36px 36px 28px;border-radius:16px;margin-bottom:36px;}}
    .header .kicker{{font-size:11px;letter-spacing:.2em;text-transform:uppercase;color:var(--blue-mid);font-weight:700;margin-bottom:8px;}}
    .header h1{{font-size:30px;font-weight:700;line-height:1.15;margin-bottom:6px;}}
    .header .subtitle{{font-size:15px;color:#a8b0c0;line-height:1.5;max-width:700px;}}
    .header .meta{{margin-top:16px;font-size:12px;color:#8090a8;border-top:1px solid #334;padding-top:12px;}}
    .summary-box{{background:var(--panel);border:1px solid var(--border);border-radius:14px;padding:28px 30px;margin-bottom:32px;box-shadow:var(--shadow);}}
    .summary-box h2{{font-size:17px;color:var(--blue);text-transform:uppercase;letter-spacing:.1em;margin-bottom:12px;font-weight:700;}}
    .summary-box p{{font-size:15.5px;color:var(--body);line-height:1.72;margin:0 0 12px;}}
    .summary-box p:last-child{{margin-bottom:0;}}
    .toc{{background:var(--panel);border:1px solid var(--border);border-radius:14px;padding:22px 24px;margin-bottom:32px;box-shadow:var(--shadow);}}
    .toc h2{{font-size:17px;color:var(--blue);text-transform:uppercase;letter-spacing:.1em;margin-bottom:12px;font-weight:700;}}
    .toc-grid{{display:grid;grid-template-columns:1fr 1fr;gap:8px 22px;}}
    .toc a{{color:var(--ink);text-decoration:none;font-size:14px;}}
    .toc a:hover{{color:var(--blue);text-decoration:underline;}}
    .kpi-strip{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:36px;}}
    .kpi{{background:var(--panel);border:1px solid var(--border);border-radius:12px;padding:18px 14px;text-align:center;box-shadow:var(--shadow);}}
    .kpi .label{{font-size:10.5px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);margin-bottom:6px;}}
    .kpi .value{{font-size:28px;font-weight:700;color:var(--ink);}}
    .kpi .note{{font-size:12.5px;color:var(--muted);margin-top:4px;}}
    h2.section{{font-size:22px;color:var(--ink);margin:44px 0 6px;font-weight:700;}}
    .section-sub{{font-size:13.5px;color:var(--muted);margin-bottom:18px;}}
    h3{{font-size:17px;color:var(--ink);margin:28px 0 8px;font-weight:700;}}
    p{{margin:0 0 14px;}}
    strong{{color:var(--ink);}}
    .focus{{background:var(--panel);border:1px solid var(--border);border-radius:14px;padding:28px 28px 24px;margin:20px 0;box-shadow:var(--shadow);position:relative;overflow:hidden;}}
    .focus::before{{content:'';position:absolute;top:0;left:0;width:5px;height:100%;border-radius:14px 0 0 14px;}}
    .focus.blue::before{{background:var(--blue);}}
    .focus.green::before{{background:var(--green);}}
    .focus.warn::before{{background:var(--warn);}}
    .focus .focus-header{{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:14px;flex-wrap:wrap;gap:6px;}}
    .focus .focus-header h3{{margin:0;font-size:18px;}}
    .focus .focus-tag{{font-size:11px;text-transform:uppercase;letter-spacing:.08em;font-weight:700;padding:3px 10px;border-radius:6px;white-space:nowrap;}}
    .focus .focus-tag.ops{{background:var(--blue-soft);color:var(--blue);}}
    .focus .focus-tag.def{{background:var(--red-bg);color:var(--red);}}
    .focus .focus-tag.pr{{background:var(--warn-bg);color:var(--warn);}}
    .insight-row{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:16px 0;}}
    .insight-stat{{background:var(--bg);border-radius:10px;padding:14px 16px;text-align:center;}}
    .insight-stat .big{{font-size:26px;font-weight:700;color:var(--ink);}}
    .insight-stat .desc{{font-size:12px;color:var(--muted);margin-top:2px;}}
    .rec{{border-left:4px solid var(--green);background:var(--green-bg);padding:16px 20px;border-radius:0 10px 10px 0;margin:16px 0 8px;}}
    .rec p{{margin:0 0 6px;}} .rec p:last-child{{margin:0;}} .rec strong{{color:var(--green);}}
    table{{width:100%;border-collapse:collapse;margin:16px 0 18px;font-size:13.5px;}}
    th,td{{padding:8px 10px;border-bottom:1px solid var(--border);text-align:left;}}
    th{{font-size:10.5px;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);border-bottom:2px solid var(--border);}}
    td.r{{text-align:right;font-variant-numeric:tabular-nums;}}
    tr.total{{font-weight:700;}} tr.total td{{border-top:2px solid var(--ink);}}
    .down{{color:var(--green);}} .up{{color:var(--red);}}
    .callout{{background:var(--purple-bg);border-left:4px solid var(--purple);padding:18px 22px;border-radius:0 10px 10px 0;margin:20px 0;}}
    .callout p:last-child{{margin:0;}}
    hr{{border:none;border-top:1px solid var(--border);margin:36px 0;}}
    .footer{{margin-top:48px;padding-top:16px;border-top:2px solid var(--ink);font-size:11.5px;color:var(--muted);text-align:center;}}
    @media print{{body{{background:#fff;font-size:11pt;}} .wrap{{max-width:100%;padding:0;}} .header{{border-radius:0;}}}}
    @media(max-width:700px){{.wrap{{padding:20px 14px 40px;}} .kpi-strip{{grid-template-columns:1fr 1fr;}} .insight-row{{grid-template-columns:1fr;}} .header h1{{font-size:24px;}}}}
  </style>
</head>
<body>
<div class="wrap">

  <div class="header">
    <div class="kicker">ORBIT Program &bull; HR Workload Lens</div>
    <h1>{CUSTOM_REPORT_LABEL} Insights</h1>
    <div class="subtitle">What changed versus Week 9, what is structurally driving the queue, and where leadership action should focus first.</div>
    <div class="meta">
      Prepared for HR Operations Leadership &nbsp;|&nbsp; {report_date}<br>
      Reporting week: {CUSTOM_REPORT_LABEL} ({curr_lbl}, 2026) &nbsp;|&nbsp; WoW reference: Week 9 ({prior_lbl}, 2026)<br>
      Scope: {' &bull; '.join(SERVICES)}
    </div>
  </div>

  <div class="summary-box">
    <h2>Focused Summary</h2>
    <p><strong>Scope note:</strong> this report analyzes 4 services and explains {enterprise_slice_pct:.1f}% of total Week 10 ticket volume ({curr_n:,} of {CUSTOM_REPORT_TOTAL_WEEK10_VOLUME:,}). It is a meaningful slice of the work, but not the full enterprise picture.</p>
    <p><strong>Data reviewed:</strong> we looked at {total_n:,} ticket records across Weeks 9 and 10 and {CUSTOM_REPORT_UKG_ACTIONS_PER_WEEK:,} UKG timecard touches as operational context, focusing on 4 of the top 5 HR service drivers by volume. LOAA remains intentionally out of scope at this stage.</p>
    <p><strong>OBR Question 1 | What is causing the week-to-week variance?</strong> {CUSTOM_REPORT_LABEL} closed at {curr_n:,} tickets across four HR services, down {wow_abs:,} ({wow_pct:.1f}%) from Week 9. The drop was led by Attendance Inquiry ({attendance_current:,}, down {abs(attendance_delta):,}) and FC General Inquiry ({fc_current:,}, down {abs(fc_delta):,}), while CC Time and Attendance remained essentially flat at {cc_current:,}. FC General Inquiry is explainable in validated data, led by lower Onboarding/I9 processing and fewer Smartsheet notifications. Attendance Inquiry is down as well, but validated process-required and defect categories explain only {attendance_validated_drop:,} of that decline, so we should not treat the full improvement as a confirmed behavior shift yet.</p>
    <p><strong>OBR Question 2 | What specific types of issues are driving variability?</strong> The validated variance is coming primarily from movement inside process-required and defect categories, not from NICE-to-UKG sync work, which stayed nearly flat at {po_current:,} this week versus {po_prior:,} last week. The largest decreases were {negative_issue_summary_topline}. The main increases were {positive_issue_summary}.</p>
    <p><strong>OBR Question 3 | What actions can stabilize the trend?</strong> AI insight and recommendation: fix the NICE-to-UKG handoff, retire Smartsheet notifications, remediate the {top_def_name.lower()} defect, and focus queue cleanup first on the five assignment groups generating {top_sites_pct} of Week 10 volume. The goal is fewer tickets of the right kind: less recurring operational rework, fewer defect-driven corrections, and a cleaner remaining queue that reflects legitimate HR work. For Attendance Inquiry, complete validation of the remaining mix before treating this week&rsquo;s drop as durable improvement or baking it into the narrative.</p>
    <p><strong>Model note:</strong> self-service candidate metrics are intentionally withheld this week while model training and validation are completed. They will be added next week once the output has been reviewed.</p>
  </div>

  <div class="toc">
    <h2>Table of Contents</h2>
    <div class="toc-grid">
      <a href="#classification-breakdown">What We Found: The Classification Breakdown</a>
      <a href="#service-mix">Week 10 Service Mix</a>
      <a href="#focus-1">Focus Area 1: NICE &amp; UKG Sync Work</a>
      <a href="#focus-2">Focus Area 2: Legitimate HR Process Volume</a>
      <a href="#focus-3">Focus Area 3: Defect Concentration</a>
      <a href="#focus-4">Focus Area 4: Where Volume Is Concentrated</a>
      <a href="#bigger-picture">The Bigger Picture: What Tickets Don&rsquo;t Show</a>
      <a href="#combined-impact">Combined Impact: What This Gets Us</a>
      <a href="#bottom-line">The Bottom Line</a>
    </div>
  </div>

  <div class="kpi-strip">
    <div class="kpi"><div class="label">Week 10 Ticket Volume</div><div class="value">{curr_n:,}</div><div class="note">{curr_lbl}</div></div>
    <div class="kpi"><div class="label">WoW Volume Change</div><div class="value" style="color:var(--{'green' if wow_delta < 0 else 'red'})">{wow_str}</div><div class="note">{prior_n:,} &rarr; {curr_n:,}</div></div>
    <div class="kpi"><div class="label">Largest Service</div><div class="value">{largest_service_current:,}</div><div class="note">{largest_service} ({largest_service_pct:.1f}%)</div></div>
    <div class="kpi"><div class="label">WFM Sync Load</div><div class="value">{po_current:,}</div><div class="note">{po_pct_current:.1f}% of Week 10</div></div>
  </div>

  <h2 class="section" id="classification-breakdown">What We Found: The Classification Breakdown</h2>
  <p class="section-sub">This table reports Week 10 only. Week 9 is shown only for change context. Self-service metrics are intentionally withheld until model validation is complete.</p>

  <table>
    <thead><tr><th>Classification</th><th>Week 10 Tickets</th><th>% of Week 10</th><th>Week 9</th><th>WoW Change</th></tr></thead>
    <tbody>{class_rows}</tbody>
  </table>

  <div class="callout">
    <p><strong>Model status:</strong> the self-service candidate model is still in training and validation. It will be added back to this section next week after the output has been reviewed and signed off.</p>
  </div>

  <h2 class="section" id="service-mix">Week 10 Service Mix</h2>
  <p class="section-sub">Attendance Inquiry still represented nearly half of total volume, while CC Time and Attendance held flat week over week.</p>

  <table>
    <thead><tr><th>HR Service</th><th>Week 10 Tickets</th><th>% of Week 10</th><th>Week 9</th><th>WoW Change</th></tr></thead>
    <tbody>{service_rows}</tbody>
  </table>

  <hr>

  <h2 class="section" id="focus-1">Focus Area 1: NICE &harr; UKG Sync Work</h2>
  <p class="section-sub">A large share of Week 10 volume is operational rework created by system handoff gaps.</p>

  <div class="focus blue">
    <div class="focus-header">
      <h3>WFM-to-HRSS Sync Requests Remain a Structural Load</h3>
      <span class="focus-tag ops">Validated</span>
    </div>
    <div class="insight-row">
      <div class="insight-stat"><div class="big">{po_current:,}</div><div class="desc">Week 10 sync tickets<br>({po_pct_current:.1f}% of all Week 10 volume)</div></div>
      <div class="insight-stat"><div class="big">{po_current - po_prior:+,}</div><div class="desc">WoW change vs Week 9<br>{po_prior:,} to {po_current:,}</div></div>
    </div>
    <p>These tickets are generated by the Real Time Analyst workflow when NICE and UKG do not align and HRSS must manually update the record. This is not normal employee demand. It is recurring operational rework created by a system handoff gap.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> treat NICE-to-UKG sync as a system stabilization priority, not an HRSS productivity issue.</p>
      <p><strong>Expected impact:</strong> resolving this handoff would remove roughly {po_current:,} tickets per week from the queue at the current run rate.</p>
    </div>
  </div>

  <hr>

  <h2 class="section" id="focus-2">Focus Area 2: Legitimate HR Process Volume</h2>
  <p class="section-sub">A material share of Week 10 demand genuinely requires HR intervention and should not be treated as avoidable noise.</p>

  <div class="focus warn">
    <div class="focus-header">
      <h3>Smartsheet and Required Process Work Are Still Material Drivers</h3>
      <span class="focus-tag pr">Validated</span>
    </div>
    <div class="insight-row">
      <div class="insight-stat"><div class="big">{pr_current_total:,}</div><div class="desc">Week 10 process-required tickets<br>({pr_current_total/curr_n*100:.1f}% of total volume)</div></div>
      <div class="insight-stat"><div class="big">{smartsheet_current:,}</div><div class="desc">Smartsheet-driven tickets in Week 10<br>({smartsheet_pct} of process-required volume)</div></div>
    </div>
    <table>
      <thead><tr><th>Process-Required Category</th><th>Week 10 Tickets</th><th>% of PR</th><th>Week 9</th><th>WoW Change</th></tr></thead>
      <tbody>{pr_rows}</tbody>
    </table>
    <p>Required HR work remains a large part of the queue, but some of it is more stabilizable than it looks. Smartsheet-driven notifications alone accounted for {smartsheet_current:,} tickets this week, and site-level examples such as fewer AVP4 Smartsheet triggers show how local process design can materially move volume.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> continue Smartsheet retirement and clean up local trigger-based workflows before treating week-to-week improvement as a pure demand story.</p>
      <p><strong>Expected impact:</strong> removing or redesigning these triggers lowers volume without reducing service quality and leaves a cleaner queue made up of the right kinds of HR work.</p>
    </div>
  </div>

  <hr>

  <h2 class="section" id="focus-3">Focus Area 3: Defect Concentration</h2>
  <p class="section-sub">The largest defect is concentrated enough to justify a direct fix rather than ongoing manual handling.</p>

  <div class="focus green">
    <div class="focus-header">
      <h3>{top_def_name}</h3>
      <span class="focus-tag def">Validated</span>
    </div>
    <div class="insight-row">
      <div class="insight-stat"><div class="big">{top_def_current:,}</div><div class="desc">Week 10 tickets<br>({top_def_pct_current} of all defects)</div></div>
      <div class="insight-stat"><div class="big">{def_current_total:,}</div><div class="desc">Total defect tickets in Week 10<br>({def_current_total/curr_n*100:.1f}% of total volume)</div></div>
    </div>
    <p>{top_def_name} remains the dominant confirmed defect in the queue. This should be treated as a fixable system issue, not as normal operating demand.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> assign an owner to audit the coding and reimbursement logic behind this defect and remove the repeat correction work.</p>
      <p><strong>Expected impact:</strong> sustaining this improvement lowers corrective tickets directly and reduces one of the clearest trust-eroding issues in the current queue.</p>
    </div>
  </div>

  <hr>

  <h2 class="section" id="focus-4">Focus Area 4: Where Volume Is Concentrated</h2>
  <p class="section-sub">Just a handful of assignment groups account for most of the Week 10 queue.</p>

  <div class="focus warn">
    <div class="focus-header">
      <h3>Top 5 Assignment Groups = {top_sites_pct} of Week 10 Volume</h3>
      <span class="focus-tag pr">Validated</span>
    </div>
    <table>
      <thead><tr><th>Assignment Group</th><th>Week 10 Tickets</th><th>% of Week 10</th><th>Week 9</th><th>WoW Change</th></tr></thead>
      <tbody>{top_sites_rows}</tbody>
    </table>
    <p><strong>{top_sites_pct} of Week 10 tickets came from just five assignment groups.</strong> That concentration makes piloting changes at a small number of queues materially more effective than broad network-wide messaging.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> target operational fixes and intake cleanup at the highest-volume assignment groups first, starting with Real Time Analyst Tier I and the largest site HR queues.</p>
      <p><strong>Expected impact:</strong> working the top five queues first gives us coverage across {top_sites_pct} of this ticket slice and the fastest signal on which interventions actually reduce demand.</p>
    </div>
  </div>

  <hr>

  <h2 class="section" id="bigger-picture">The Bigger Picture: What Tickets Don&rsquo;t Show</h2>
  <p class="section-sub">Tickets are only the visible layer of the work. The underlying operational load is materially larger.</p>

  <p>This report explains a meaningful slice of visible demand, but not the full workload picture. These 4 services represent {enterprise_slice_pct:.1f}% of total Week 10 ticket volume ({curr_n:,} of {CUSTOM_REPORT_TOTAL_WEEK10_VOLUME:,}), and ticket counts alone do not capture the proactive HR and WFM work happening underneath them.</p>

  <table>
    <thead><tr><th>Workload Signal</th><th>Weekly Volume</th><th>Implication</th></tr></thead>
    <tbody>
      <tr><td>Tickets in this report</td><td class="r">{curr_n:,}</td><td>This is the visible slice we are explaining here</td></tr>
      <tr><td>Total Week 10 ticket volume</td><td class="r">{CUSTOM_REPORT_TOTAL_WEEK10_VOLUME:,}</td><td>Our current scope covers about {enterprise_slice_pct:.1f}% of the full enterprise queue</td></tr>
      <tr><td>UKG actions per week</td><td class="r">{CUSTOM_REPORT_UKG_ACTIONS_PER_WEEK:,}</td><td>Underlying operational activity is far larger than ticket counts alone suggest</td></tr>
      <tr><td>UKG rework actions per week</td><td class="r">{CUSTOM_REPORT_UKG_REWORK_ACTIONS_PER_WEEK:,}</td><td>Ticket volume maps to only {visible_vs_rework_pct:.1f}% of known rework activity</td></tr>
      <tr><td>Estimated FTE hours per week</td><td class="r">{CUSTOM_REPORT_UKG_FTE_HOURS_PER_WEEK:,}</td><td>Workload remains meaningful even when tickets improve WoW</td></tr>
    </tbody>
  </table>

  <div class="callout">
    <p><strong>Model direction:</strong> this model will increasingly look across both ticket signals and operational workload signals so future readouts explain not only what changed in tickets, but what changed in the actual work underneath them.</p>
  </div>

  <hr>

  <h2 class="section" id="combined-impact">Combined Impact: What This Gets Us</h2>

  <table>
    <thead><tr><th>Focus Area</th><th>Current Signal</th><th>What Stabilizing It Gets Us</th></tr></thead>
    <tbody>
      <tr><td>NICE &harr; UKG handoff</td><td class="r">{po_current:,} tickets / week</td><td>Less recurring operational rework and a more stable baseline</td></tr>
      <tr><td>Smartsheet and trigger cleanup</td><td class="r">{smartsheet_current:,} tickets / week</td><td>Fewer auto-generated tickets and less site-specific noise</td></tr>
      <tr><td>Time-off reimbursement defect</td><td class="r">{top_def_current:,} tickets / week</td><td>Fewer corrective contacts and better trust in the system</td></tr>
      <tr><td>Top 5 assignment groups</td><td class="r">{top_sites_pct} of this slice</td><td>Faster interventions and clearer signal on what actually works</td></tr>
      <tr class="total"><td colspan="2"><strong>Combined effect</strong></td><td><strong>Fewer tickets of the right kind: less rework and defect noise, with a remaining queue that is easier to explain and stabilize</strong></td></tr>
    </tbody>
  </table>

  <p>That is the immediate value of these four focus areas. The next step is better precision: as model validation continues, we will be able to move from high-level stabilization insights to deeper root-cause guidance and sharper recommendations.</p>

  <hr>

  <h2 class="section" id="bottom-line">The Bottom Line</h2>

  <p>Week 10 improved versus Week 9, but this report shows why stabilization is not just about fewer tickets. It is about fewer tickets of the right kind: less NICE-to-UKG rework, fewer trigger-driven process tickets, fewer defect corrections, and better concentration of intervention where the queue is heaviest. This week&rsquo;s report intentionally stays inside the bounds of what has been validated. Self-service candidate metrics will be added next week once the model review is complete.</p>

  <hr>

  <p style="font-size:13px;color:var(--muted);line-height:1.5;">
    <strong>Methodology:</strong> Reported week is {CUSTOM_REPORT_LABEL} ({curr_lbl}, 2026) across 4 HR services representing {enterprise_slice_pct:.1f}% of total Week 10 ticket volume ({curr_n:,} of {CUSTOM_REPORT_TOTAL_WEEK10_VOLUME:,}). Week 9 ({prior_lbl}, 2026) is used only as change context. Classification is rule-based on ticket description and contact type, but self-service candidate metrics are withheld in this version while model training and validation are in progress. Separate UKG workload figures are included as directional operational context. Report generated automatically by the ORBIT Workload Lens pipeline.
  </p>

  <div class="footer">
    ORBIT HR Workload Lens &nbsp;|&nbsp; {CUSTOM_REPORT_LABEL} Insights &nbsp;|&nbsp; {curr_lbl}, 2026<br>
    Generated {report_date} &nbsp;|&nbsp; Pipeline v1.0
  </div>

</div>
</body>
</html>"""

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return output_path


def generate_report(metrics, obr_text, output_path):
    """Generate the HTML insights report."""
    if metrics['current_start'].strftime('%Y-%m-%d') == CUSTOM_REPORT_WEEK:
        return generate_custom_week_report(metrics, output_path)

    m = metrics
    prior_n = m['total_prior']
    curr_n = m['total_current']
    total = m['total_all']
    wow_delta, wow_pct, wow_str, wow_color = delta_str(prior_n, curr_n)

    # Classification rows
    class_order = ['Self Service Eligible', 'Process Optimization', 'Process Required', 'Defect', 'Unclear']
    ss_total = m['class_counts'].get('Self Service Eligible', {}).get('total', 0)
    ss_pct_of_total = ss_total / total * 100 if total else 0
    assumed_ss_pct_of_total = SELF_SERVICE_POTENTIAL_ASSUMPTION * 100 if total else 0
    assumed_ss_total = round(total * SELF_SERVICE_POTENTIAL_ASSUMPTION)
    additional_ss_potential = max(0, assumed_ss_total - ss_total)

    # Top SS sub-categories
    ss_sorted = sorted(m['ss_subcats'].items(),
                       key=lambda x: x[1].get('prior', 0) + x[1].get('current', 0),
                       reverse=True)

    # Top PR sub-categories
    pr_sorted = sorted(m['pr_subcats'].items(),
                       key=lambda x: x[1].get('prior', 0) + x[1].get('current', 0),
                       reverse=True)

    # Top Defect sub-categories
    def_sorted = sorted(m['def_subcats'].items(),
                        key=lambda x: x[1].get('prior', 0) + x[1].get('current', 0),
                        reverse=True)

    # Top sites
    site_sorted = sorted(m['site_counts'].items(),
                         key=lambda x: x[1].get('prior', 0) + x[1].get('current', 0),
                         reverse=True)

    # Week labels
    prior_lbl = m['prior_label']
    curr_lbl = m['current_label']
    prior_wk_num = m['prior_start'].isocalendar()[1]
    curr_wk_num = m['current_start'].isocalendar()[1]
    report_date = datetime.now().strftime('%B %d, %Y')

    # ── Build HTML ──
    # Helper to make table rows
    def class_row(name, data, is_total=False):
        t = data.get('total', 0)
        p = data.get('prior', 0)
        c = data.get('current', 0)
        d, dp, ds, dc = delta_str(p, c)
        bold = "<strong>" if is_total or name == "Self Service Eligible" else ""
        bold_e = "</strong>" if bold else ""
        tr_class = ' class="total"' if is_total else ''
        pct_t = f"{t/total*100:.1f}%" if total else "0%"
        return f"""<tr{tr_class}><td>{bold}{name}{bold_e}</td><td class="r">{bold}{t:,}{bold_e}</td><td class="r">{bold}{pct_t}{bold_e}</td><td class="r">{p:,}</td><td class="r">{c:,}</td><td class="r {dc}">{ds}</td></tr>"""

    class_rows = ""
    for cn in class_order:
        data = m['class_counts'].get(cn, {'total': 0, 'prior': 0, 'current': 0})
        class_rows += class_row(cn, data)
    class_rows += class_row("Total", {'total': total, 'prior': prior_n, 'current': curr_n}, True)

    # Service rows
    svc_rows = ""
    for svc in SERVICES:
        data = m['svc_counts'].get(svc, {'prior': 0, 'current': 0})
        p, c = data.get('prior', 0), data.get('current', 0)
        t = p + c
        d, dp, ds, dc = delta_str(p, c)
        pct_t = f"{t/total*100:.1f}%" if total else "0%"
        svc_rows += f'<tr><td>{svc}</td><td class="r">{p:,}</td><td class="r">{c:,}</td><td class="r {dc}">{ds}</td><td class="r">{pct_t}</td></tr>'
    svc_rows += f'<tr class="total"><td>Total</td><td class="r">{prior_n:,}</td><td class="r">{curr_n:,}</td><td class="r {wow_color}">{wow_str}</td><td class="r">100%</td></tr>'

    # SS sub-category rows (top 10 + other)
    ss_rows = ""
    ss_total_val = sum(v.get('prior', 0) + v.get('current', 0) for v in m['ss_subcats'].values())
    shown = 0
    other_p, other_c = 0, 0
    for i, (sc, data) in enumerate(ss_sorted):
        p, c = data.get('prior', 0), data.get('current', 0)
        t = p + c
        if i < 7:
            pct_ss = f"{t/ss_total_val*100:.1f}%" if ss_total_val else "0%"
            ch = data.get('channel', '')
            bold = "<strong>" if i < 2 else ""
            bold_e = "</strong>" if bold else ""
            ss_rows += f'<tr><td>{bold}{sc}{bold_e}</td><td class="r">{p:,}</td><td class="r">{c:,}</td><td class="r">{bold}{t:,}{bold_e}</td><td class="r">{bold}{pct_ss}{bold_e}</td><td>{ch}</td></tr>'
            shown += t
        else:
            other_p += p
            other_c += c
    if other_p + other_c > 0:
        ot = other_p + other_c
        pct_ot = f"{ot/ss_total_val*100:.1f}%" if ss_total_val else "0%"
        ss_rows += f'<tr><td>All other self-service categories</td><td class="r">{other_p:,}</td><td class="r">{other_c:,}</td><td class="r">{ot:,}</td><td class="r">{pct_ot}</td><td>Various</td></tr>'
    ss_rows += f'<tr class="total"><td>Total Self Service Eligible</td><td class="r">{m["class_counts"].get("Self Service Eligible", {}).get("prior", 0):,}</td><td class="r">{m["class_counts"].get("Self Service Eligible", {}).get("current", 0):,}</td><td class="r">{ss_total_val:,}</td><td class="r">100%</td><td></td></tr>'

    # PR sub-category rows (top 10)
    pr_rows = ""
    pr_total_val = sum(v.get('prior', 0) + v.get('current', 0) for v in m['pr_subcats'].values())
    pr_shown_p, pr_shown_c, pr_shown_t = 0, 0, 0
    for i, (sc, data) in enumerate(pr_sorted):
        p, c = data.get('prior', 0), data.get('current', 0)
        t = p + c
        if i < 9:
            pct_pr = f"{t/pr_total_val*100:.1f}%" if pr_total_val else "0%"
            pr_rows += f'<tr><td>{sc}</td><td class="r">{p:,}</td><td class="r">{c:,}</td><td class="r">{t:,}</td><td class="r">{pct_pr}</td></tr>'
            pr_shown_p += p; pr_shown_c += c; pr_shown_t += t
        else:
            break
    oth_p = m['class_counts'].get('Process Required', {}).get('prior', 0) - pr_shown_p
    oth_c = m['class_counts'].get('Process Required', {}).get('current', 0) - pr_shown_c
    if oth_p + oth_c > 0:
        ot = oth_p + oth_c
        pr_rows += f'<tr><td>All other process-required</td><td class="r">{oth_p:,}</td><td class="r">{oth_c:,}</td><td class="r">{ot:,}</td><td class="r">{ot/pr_total_val*100:.1f}%</td></tr>'
    pr_rows += f'<tr class="total"><td>Total Process Required</td><td class="r">{m["class_counts"].get("Process Required", {}).get("prior", 0):,}</td><td class="r">{m["class_counts"].get("Process Required", {}).get("current", 0):,}</td><td class="r">{pr_total_val:,}</td><td class="r">100%</td></tr>'

    # Top sites table (top 5)
    site_rows_top5 = ""
    running_total = 0
    for i, (ag, data) in enumerate(site_sorted[:5]):
        p, c = data.get('prior', 0), data.get('current', 0)
        t = p + c
        running_total += t
        ss_p = data.get('ss_prior', 0)
        ss_c = data.get('ss_current', 0)
        ss_t = ss_p + ss_c
        ss_rate = f"{ss_t/t*100:.0f}%" if t else "—"
        d, dp, ds, dc = delta_str(p, c)
        pct_all = f"{t/total*100:.1f}%" if total else "0%"
        site_rows_top5 += f'<tr><td><strong>{ag}</strong></td><td class="r"><strong>{t:,}</strong></td><td class="r"><strong>{pct_all}</strong></td><td class="r">{ss_rate}</td><td class="r">{p:,}</td><td class="r">{c:,}</td><td class="r {dc}">{ds}</td></tr>'
    top5_pct = f"{running_total/total*100:.1f}%" if total else "0%"
    site_rows_top5 += f'<tr class="total"><td>Top 5 Subtotal</td><td class="r">{running_total:,}</td><td class="r">{top5_pct}</td><td class="r"></td><td class="r"></td><td class="r"></td><td class="r"></td></tr>'

    # High SS-rate sites (>= 80% SS, >= 50 tickets)
    high_ss_sites = []
    for ag, data in site_sorted:
        t = data.get('prior', 0) + data.get('current', 0)
        ss_t = data.get('ss_prior', 0) + data.get('ss_current', 0)
        if t >= 50 and ss_t / t >= 0.78:
            d, dp, ds, dc = delta_str(data.get('prior', 0), data.get('current', 0))
            high_ss_sites.append((ag, t, ss_t/t*100, ds))

    high_ss_rows = ""
    for ag, t, ss_rate, ds in high_ss_sites[:5]:
        high_ss_rows += f'<tr><td>{ag}</td><td class="r">{t:,}</td><td class="r"><strong>{ss_rate:.0f}%</strong></td><td class="r">{ds}</td></tr>'

    # Top defect
    top_defect = def_sorted[0] if def_sorted else ("None", {'prior': 0, 'current': 0})
    top_def_name = top_defect[0]
    top_def_total = top_defect[1].get('prior', 0) + top_defect[1].get('current', 0)
    def_total_val = sum(v.get('prior', 0) + v.get('current', 0) for v in m['def_subcats'].values())
    top_def_pct = f"{top_def_total/def_total_val*100:.1f}%" if def_total_val else "0%"

    # Service x Classification for the narrative
    svc_class_summary = {}
    for svc in SERVICES:
        sc = m['svc_class'].get(svc, {})
        svc_total = sum(
            v.get('prior', 0) + v.get('current', 0)
            for v in sc.values()
        )
        ss_svc = sc.get('Self Service Eligible', {})
        ss_svc_t = ss_svc.get('prior', 0) + ss_svc.get('current', 0)
        def_svc = sc.get('Defect', {})
        def_svc_t = def_svc.get('prior', 0) + def_svc.get('current', 0)
        pr_svc = sc.get('Process Required', {})
        pr_svc_t = pr_svc.get('prior', 0) + pr_svc.get('current', 0)
        po_svc = sc.get('Process Optimization', {})
        po_svc_t = po_svc.get('prior', 0) + po_svc.get('current', 0)
        svc_class_summary[svc] = {
            'total': svc_total,
            'ss_pct': f"{ss_svc_t/svc_total*100:.1f}%" if svc_total else "0%",
            'def_pct': f"{def_svc_t/svc_total*100:.1f}%" if svc_total else "0%",
            'pr_pct': f"{pr_svc_t/svc_total*100:.1f}%" if svc_total else "0%",
            'po_pct': f"{po_svc_t/svc_total*100:.1f}%" if svc_total else "0%",
        }

    svc_class_rows = ""
    for svc in SERVICES:
        s = svc_class_summary[svc]
        svc_class_rows += f'<tr><td>{svc}</td><td class="r">{s["total"]:,}</td><td class="r"><strong>{s["ss_pct"]}</strong></td><td class="r">{s["po_pct"]}</td><td class="r">{s["def_pct"]}</td><td class="r">{s["pr_pct"]}</td></tr>'

    # ── Missed punch stats ──
    mp_total = sum(
        v.get('prior', 0) + v.get('current', 0)
        for k, v in m['ss_subcats'].items()
        if 'missed punch' in k.lower() or 'punch' in k.lower()
    )
    mp_prior = sum(v.get('prior', 0) for k, v in m['ss_subcats'].items() if 'missed punch' in k.lower() or 'punch' in k.lower())
    mp_current = sum(v.get('current', 0) for k, v in m['ss_subcats'].items() if 'missed punch' in k.lower() or 'punch' in k.lower())

    # ── Absence cluster stats ──
    absence_keys = ['call out', 'absence', 'voicemail', 'weather']
    abs_total = sum(
        v.get('prior', 0) + v.get('current', 0)
        for k, v in m['ss_subcats'].items()
        if any(ak in k.lower() for ak in absence_keys)
    )
    abs_prior = sum(v.get('prior', 0) for k, v in m['ss_subcats'].items() if any(ak in k.lower() for ak in absence_keys))
    abs_current = sum(v.get('current', 0) for k, v in m['ss_subcats'].items() if any(ak in k.lower() for ak in absence_keys))
    abs_d, abs_dp, abs_ds, abs_dc = delta_str(abs_prior, abs_current)

    # ── Balance inquiry stats ──
    bal_total = sum(v.get('prior', 0) + v.get('current', 0) for k, v in m['ss_subcats'].items() if 'balance' in k.lower())
    bal_prior = sum(v.get('prior', 0) for k, v in m['ss_subcats'].items() if 'balance' in k.lower())
    bal_current = sum(v.get('current', 0) for k, v in m['ss_subcats'].items() if 'balance' in k.lower())
    bal_d, bal_dp, bal_ds, bal_dc = delta_str(bal_prior, bal_current)

    # ── Smartsheet stats ──
    sm_total = sum(v.get('prior', 0) + v.get('current', 0) for k, v in m['pr_subcats'].items() if 'smartsheet' in k.lower())
    sm_pct_pr = f"{sm_total/pr_total_val*100:.1f}%" if pr_total_val else "0%"

    # ── Process Optimization stats ──
    po_total = m['class_counts'].get('Process Optimization', {}).get('total', 0)
    po_prior = m['class_counts'].get('Process Optimization', {}).get('prior', 0)
    po_current = m['class_counts'].get('Process Optimization', {}).get('current', 0)
    po_pct_of_total = po_total / total * 100 if total else 0

    # ── Now build the full HTML ──
    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>HR Workload Lens — Insights &amp; Recommendations</title>
  <style>
    :root {{--ink:#1a1f2b;--body:#3a3f4c;--muted:#717888;--border:#e2e6ee;--panel:#ffffff;--bg:#f7f8fb;--blue:#0d5f73;--blue-soft:#e4f3f7;--blue-mid:#b0dce8;--green:#14643a;--green-bg:#e6f4ea;--red:#9b2c2c;--red-bg:#fdf0f0;--warn:#7a3e00;--warn-bg:#fff7ed;--purple:#5b21b6;--purple-bg:#f3edff;--shadow:0 4px 20px rgba(20,30,50,0.06);}}
    *{{box-sizing:border-box;margin:0;}}
    body{{font-family:Aptos,"Segoe UI",Calibri,sans-serif;color:var(--body);background:var(--bg);line-height:1.68;font-size:15.5px;}}
    .wrap{{max-width:860px;margin:0 auto;padding:40px 28px 64px;}}
    .header{{background:var(--ink);color:#fff;padding:36px 36px 28px;border-radius:16px;margin-bottom:36px;}}
    .header .kicker{{font-size:11px;letter-spacing:.2em;text-transform:uppercase;color:var(--blue-mid);font-weight:700;margin-bottom:8px;}}
    .header h1{{font-size:28px;font-weight:700;line-height:1.2;margin-bottom:6px;}}
    .header .subtitle{{font-size:15px;color:#a8b0c0;line-height:1.5;}}
    .header .meta{{margin-top:16px;font-size:12px;color:#8090a8;border-top:1px solid #334;padding-top:12px;}}
    .summary-box{{background:var(--panel);border:1px solid var(--border);border-radius:14px;padding:28px 30px;margin-bottom:32px;box-shadow:var(--shadow);}}
    .summary-box h2{{font-size:17px;color:var(--blue);text-transform:uppercase;letter-spacing:.1em;margin-bottom:12px;font-weight:700;}}
    .summary-box p{{font-size:15.5px;color:var(--body);line-height:1.72;margin:0;}}
    .kpi-strip{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:36px;}}
    .kpi{{background:var(--panel);border:1px solid var(--border);border-radius:12px;padding:18px 14px;text-align:center;box-shadow:var(--shadow);}}
    .kpi .label{{font-size:10.5px;text-transform:uppercase;letter-spacing:.08em;color:var(--muted);margin-bottom:6px;}}
    .kpi .value{{font-size:28px;font-weight:700;color:var(--ink);}}
    .kpi .note{{font-size:12.5px;color:var(--muted);margin-top:4px;}}
    h2.section{{font-size:22px;color:var(--ink);margin:44px 0 6px;font-weight:700;}}
    .section-sub{{font-size:13.5px;color:var(--muted);margin-bottom:18px;}}
    h3{{font-size:17px;color:var(--ink);margin:28px 0 8px;font-weight:700;}}
    p{{margin:0 0 14px;}}
    strong{{color:var(--ink);}}
    .focus{{background:var(--panel);border:1px solid var(--border);border-radius:14px;padding:28px 28px 24px;margin:20px 0;box-shadow:var(--shadow);position:relative;overflow:hidden;}}
    .focus::before{{content:'';position:absolute;top:0;left:0;width:5px;height:100%;border-radius:14px 0 0 14px;}}
    .focus.blue::before{{background:var(--blue);}}
    .focus.green::before{{background:var(--green);}}
    .focus.purple::before{{background:var(--purple);}}
    .focus.warn::before{{background:var(--warn);}}
    .focus .focus-header{{display:flex;justify-content:space-between;align-items:baseline;margin-bottom:14px;flex-wrap:wrap;gap:6px;}}
    .focus .focus-header h3{{margin:0;font-size:18px;}}
    .focus .focus-tag{{font-size:11px;text-transform:uppercase;letter-spacing:.08em;font-weight:700;padding:3px 10px;border-radius:6px;white-space:nowrap;}}
    .focus .focus-tag.ss{{background:var(--blue-soft);color:var(--blue);}}
    .focus .focus-tag.def{{background:var(--red-bg);color:var(--red);}}
    .focus .focus-tag.pr{{background:var(--warn-bg);color:var(--warn);}}
    .insight-row{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin:16px 0;}}
    .insight-stat{{background:var(--bg);border-radius:10px;padding:14px 16px;text-align:center;}}
    .insight-stat .big{{font-size:26px;font-weight:700;color:var(--ink);}}
    .insight-stat .desc{{font-size:12px;color:var(--muted);margin-top:2px;}}
    .rec{{border-left:4px solid var(--green);background:var(--green-bg);padding:16px 20px;border-radius:0 10px 10px 0;margin:16px 0 8px;}}
    .rec p{{margin:0 0 6px;}} .rec p:last-child{{margin:0;}} .rec strong{{color:var(--green);}}
    table{{width:100%;border-collapse:collapse;margin:16px 0 18px;font-size:13.5px;}}
    th,td{{padding:8px 10px;border-bottom:1px solid var(--border);text-align:left;}}
    th{{font-size:10.5px;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);border-bottom:2px solid var(--border);}}
    td.r{{text-align:right;font-variant-numeric:tabular-nums;}}
    tr.total{{font-weight:700;}} tr.total td{{border-top:2px solid var(--ink);}}
    .down{{color:var(--green);}} .up{{color:var(--red);}}
    .callout{{background:var(--purple-bg);border-left:4px solid var(--purple);padding:18px 22px;border-radius:0 10px 10px 0;margin:20px 0;}}
    .callout p:last-child{{margin:0;}}
    hr{{border:none;border-top:1px solid var(--border);margin:36px 0;}}
    .footer{{margin-top:48px;padding-top:16px;border-top:2px solid var(--ink);font-size:11.5px;color:var(--muted);text-align:center;}}
    @media print{{body{{background:#fff;font-size:11pt;}} .wrap{{max-width:100%;padding:0;}} .header{{border-radius:0;}}}}
    @media(max-width:700px){{.wrap{{padding:20px 14px 40px;}} .kpi-strip{{grid-template-columns:1fr 1fr;}} .insight-row{{grid-template-columns:1fr;}} .header h1{{font-size:22px;}}}}
  </style>
</head>
<body>
<div class="wrap">

  <div class="header">
    <div class="kicker">ORBIT Program &bull; HR Workload Lens</div>
    <h1>What&rsquo;s Driving Our Ticket Volume &mdash; and How We Stabilize It</h1>
    <div class="subtitle">Insights and recommendations from a two-week ticket classification analysis</div>
    <div class="meta">
      Prepared for HR Operations Leadership &nbsp;|&nbsp; {report_date}<br>
      Prior Week ({prior_lbl}) &amp; Current Week ({curr_lbl})<br>
      Scope: {' &bull; '.join(SERVICES)}
    </div>
  </div>

  <div class="summary-box">
    <h2>Executive Summary</h2>
    <p>We classified all {total:,} tickets opened across four HR services during the prior and current weeks to understand what is driving the week-to-week volume swings that appear in the OBR. For this version of the report, we are using a <strong>working assumption that 80% of tickets ({assumed_ss_total:,}) could be handled through self-service</strong> once local routing artifacts such as voicemail-to-email and limited intake detail are normalized. The current rules explicitly confirm {ss_pct_of_total:.1f}% ({ss_total:,}) today, leaving another ~{additional_ss_potential:,} tickets in likely self-service potential that cannot yet be cleanly proven from the raw descriptions alone. The two biggest observed drivers are TMs not using timeclocks for punches ({mp_total:,} tickets, {mp_total/total*100:.1f}% of all volume) and TMs calling in absences instead of using available self-service tools ({abs_total:,} tickets, {abs_total/total*100:.1f}%). These two behaviors alone generate {mp_total+abs_total:,} tickets every two weeks. Total volume changed {wow_str} week over week, but the decline was concentrated in human-initiated demand. The recommendations in this report focus on four areas: timeclock compliance, app and desktop adoption for time-off and balance inquiries, targeted site campaigns at the highest-volume locations, and system fixes for the dominant defect type.</p>
  </div>

  <div class="kpi-strip">
    <div class="kpi"><div class="label">Tickets Classified</div><div class="value">{total:,}</div><div class="note">2 weeks, 4 services</div></div>
    <div class="kpi"><div class="label">Self-Service Potential</div><div class="value" style="color:var(--blue)">{assumed_ss_pct_of_total:.0f}%</div><div class="note">Observed: {ss_pct_of_total:.1f}%</div></div>
    <div class="kpi"><div class="label">WoW Volume Change</div><div class="value" style="color:var(--{'green' if wow_delta < 0 else 'red'})">{wow_str}</div><div class="note">{prior_n:,} &rarr; {curr_n:,}</div></div>
    <div class="kpi"><div class="label">Prior Week</div><div class="value">{prior_n:,}</div><div class="note">{prior_lbl}</div></div>
  </div>

  <h2 class="section">What We Found: The Classification Breakdown</h2>
  <p class="section-sub">Every ticket was reviewed and classified based on whether the TM had a self-service option available.</p>

  <table>
    <thead><tr><th>Classification</th><th>Total</th><th>% of All</th><th>Prior Wk</th><th>Current Wk</th><th>WoW Change</th></tr></thead>
    <tbody>{class_rows}</tbody>
  </table>

  <p><strong>Working assumption: {assumed_ss_pct_of_total:.0f}% of tickets &mdash; about {assumed_ss_total:,} out of {total:,} &mdash; could have been handled through existing self-service tools.</strong> The current classifier explicitly confirms {ss_total:,} of those tickets today; the remaining potential reflects incomplete intake detail and voicemail-to-email routing that obscures the true TM action.</p>

  <table>
    <thead><tr><th>HR Service</th><th>Prior Wk</th><th>Current Wk</th><th>WoW Change</th><th>% of Total</th></tr></thead>
    <tbody>{svc_rows}</tbody>
  </table>

  <table>
    <thead><tr><th>HR Service</th><th>Total</th><th>Self Service %</th><th>Process Opt %</th><th>Defect %</th><th>Process Req %</th></tr></thead>
    <tbody>{svc_class_rows}</tbody>
  </table>

  {f'''
  <div class="callout" style="border-left-color:var(--warn);background:var(--warn-bg);">
    <p><strong>Data Source Note &mdash; Process Optimization ({po_total:,} tickets, {po_pct_of_total:.1f}%)</strong></p>
    <p>These tickets originate from <strong>CC Time and Attendance</strong> and are submitted by the <strong>Real Time Analyst Tier I/II</strong> assignment group (WFM). They represent cases where WFM is asking HRSS to manually update UKG because NICE (the WFM scheduling system) and UKG are out of sync. This is <em>not</em> TM self-service and <em>not</em> a standard HR process &mdash; it is a <strong>system integration gap</strong> between NICE and UKG that creates manual rework. {po_total:,} tickets over two weeks ({po_prior:,} prior / {po_current:,} current) are classified as Process Optimization to separate them from both TM-driven volume and legitimate HR casework.</p>
    <p><strong>Recommendation:</strong> Shore up the NICE &harr; UKG integration so that schedule and time data sync automatically. Eliminating this gap would remove ~{po_total//2:,} tickets per week from the HRSS queue &mdash; a direct capacity return to the team.</p>
  </div>
  ''' if po_total > 0 else ''}

  <hr>

  <h2 class="section">Focus Area 1: Missed Punches &amp; Timeclock Compliance</h2>
  <p class="section-sub">The single largest ticket category &mdash; preventable through timeclock usage.</p>

  <div class="focus blue">
    <div class="focus-header">
      <h3>Chewtopians Are Not Using the Timeclock</h3>
      <span class="focus-tag ss">Self Service Eligible</span>
    </div>
    <div class="insight-row">
      <div class="insight-stat"><div class="big">{mp_total:,}</div><div class="desc">Missed punch tickets in 2 weeks<br>({mp_total/total*100:.1f}% of all tickets)</div></div>
      <div class="insight-stat"><div class="big">{mp_prior:,} / {mp_current:,}</div><div class="desc">Prior Wk / Current Wk<br>Stable &mdash; this is structural, not a spike</div></div>
    </div>
    <p>Missed punches are the #1 ticket category. When a TM fails to clock in or out, they contact HR to fix it. The timeclock was available. This category is remarkably stable week over week ({mp_prior:,} vs. {mp_current:,}), meaning it is not driving the variance &mdash; it is the <strong>structural floor</strong> underneath the variance. It makes the baseline high, so that every fluctuation in other categories looks proportionally larger.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> Launch a targeted <strong>timeclock compliance communication campaign</strong> at the top-volume sites. Reinforce timeclock usage through shift-start huddle reminders, visible signage near entry points, and manager accountability. For CC/CVC locations, promote <strong>desktop UKG access</strong> for clocking in/out at workstation.</p>
      <p><strong>Expected impact:</strong> A 25% improvement would eliminate ~{mp_total//8:,} tickets/week.</p>
    </div>
  </div>

  <hr>

  <h2 class="section">Focus Area 2: Absence Reporting &amp; Call-Off Intake</h2>
  <p class="section-sub">The #1 variance driver. When TMs call in, they call HR instead of using the app.</p>

  <div class="focus blue">
    <div class="focus-header">
      <h3>Phone Calls and Voicemails Are Driving the Weekly Swing</h3>
      <span class="focus-tag ss">Self Service Eligible</span>
    </div>
    <div class="insight-row">
      <div class="insight-stat"><div class="big">{abs_total:,}</div><div class="desc">Absence / call-out tickets in 2 weeks<br>({abs_total/total*100:.1f}% of all tickets)</div></div>
      <div class="insight-stat"><div class="big">{abs_ds}</div><div class="desc">WoW change in absence reporting<br>The #1 variance driver</div></div>
    </div>
    <p>When a Chewtopian needs to report an absence, request time off, or call out sick, the UKG mobile app allows them to do this directly. Instead, {abs_total:,} TMs over two weeks picked up the phone, sent an email, or left a voicemail for HR. This category swings the most week to week, making it the primary source of the variance leadership sees in the OBR.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> Drive adoption of the <strong>UKG mobile app</strong> for absence and call-off submissions. For Chewtopians who prefer desktop, promote <strong>desktop UKG access</strong> for submitting time-off requests. Pair with a <strong>communication campaign</strong>: <em>&ldquo;Need to call out? Use the UKG app. Need to request PTO? Use UKG or Workday. No phone call necessary.&rdquo;</em></p>
      <p><strong>Expected impact:</strong> A 30% adoption shift would deflect ~{abs_total*3//20:,} tickets per week.</p>
    </div>
  </div>

  <hr>

  <h2 class="section">Focus Area 3: Time-Off Balances &amp; Self-Service Visibility</h2>
  <p class="section-sub">The fastest-growing inquiry type &mdash; TMs are calling HR to ask questions they can answer themselves.</p>

  <div class="focus purple">
    <div class="focus-header">
      <h3>Chewtopians Cannot Easily See Their Own Balances</h3>
      <span class="focus-tag ss">Self Service Eligible</span>
    </div>
    <div class="insight-row">
      <div class="insight-stat"><div class="big">{bal_ds}</div><div class="desc">WoW change in balance inquiries<br>Fastest-growing self-service category</div></div>
      <div class="insight-stat"><div class="big">{bal_total:,}</div><div class="desc">Balance inquiry tickets in 2 weeks</div></div>
    </div>
    <p>PTO/UTO balance inquiries are growing. TMs are checking their balances through HR rather than UKG or Workday, where the information is already available. The root cause is visibility &mdash; TMs either don't know where to look or don't trust what they see.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> Promote <strong>UKG app and desktop UKG access</strong> as the primary channels for balance checks and time-off submissions. Post QR codes at break rooms and timeclocks: <em>&ldquo;Check your PTO/UTO balance anytime &mdash; no need to call HR.&rdquo;</em> For CC/CVC, bookmark desktop UKG on workstations with a one-click path to balance lookup.</p>
    </div>
  </div>

  <hr>

  <h2 class="section">Focus Area 4: Concentrated Volume &mdash; Where to Target First</h2>
  <p class="section-sub">A small number of sites generate a disproportionate share of volume. Targeted campaigns here will have outsized impact.</p>

  <div class="focus warn">
    <div class="focus-header">
      <h3>Top 5 Assignment Groups = {top5_pct} of All Tickets</h3>
      <span class="focus-tag pr">Targeted Opportunity</span>
    </div>
    <table>
      <thead><tr><th>Assignment Group</th><th>Total</th><th>% of All</th><th>SS %</th><th>Prior</th><th>Current</th><th>WoW Change</th></tr></thead>
      <tbody>{site_rows_top5}</tbody>
    </table>
    <p><strong>Just five assignment groups account for {top5_pct} of all tickets.</strong> Self-service rates at these locations are high, meaning a targeted adoption campaign would have outsized impact.</p>
    {"<p>Sites with the highest self-service rates (&ge;78%, &ge;50 tickets) represent quick-win campaign targets:</p><table><thead><tr><th>Site</th><th>Total</th><th>SS %</th><th>WoW Change</th></tr></thead><tbody>" + high_ss_rows + "</tbody></table>" if high_ss_rows else ""}
    <div class="rec">
      <p><strong>Recommendation:</strong> Run <strong>targeted self-service adoption campaigns</strong> at the highest-volume locations first. Focus on timeclock compliance (FC sites) and desktop UKG access (CC). Then expand to high-SS-rate sites where the campaign message is simple and deflection potential is highest.</p>
    </div>
  </div>

  <hr>

  <h2 class="section">Additional Insight: The Dominant Defect</h2>
  <p class="section-sub">A single system issue generates the majority of defect tickets.</p>

  <div class="focus green">
    <div class="focus-header">
      <h3>{top_def_name}</h3>
      <span class="focus-tag def">System Defect</span>
    </div>
    <div class="insight-row">
      <div class="insight-stat"><div class="big">{top_def_total:,}</div><div class="desc">{top_def_name} tickets<br>({top_def_pct} of all defects)</div></div>
      <div class="insight-stat"><div class="big">{def_total_val:,}</div><div class="desc">Total defect tickets in 2 weeks<br>({def_total_val/total*100:.1f}% of all volume)</div></div>
    </div>
    <p>Defects are a small share of total volume ({def_total_val/total*100:.1f}%), but <strong>{top_def_name} accounts for {top_def_pct} of all defects</strong>. This erodes TM trust in the system and drives more calls to HR instead of self-service usage.</p>
    <div class="rec">
      <p><strong>Recommendation:</strong> Engage the system integration team to audit the coding logic that generates these errors. This is a <strong>system configuration fix</strong>, not a people fix.</p>
    </div>
  </div>

  <hr>

  <h2 class="section">Process Required: What Legitimately Needs HR</h2>
  <p class="section-sub">{m['class_counts'].get('Process Required', {}).get('total', 0)/total*100:.1f}% of tickets require genuine HR involvement.</p>

  <table>
    <thead><tr><th>Sub-Category</th><th>Prior</th><th>Current</th><th>Total</th><th>% of PR</th></tr></thead>
    <tbody>{pr_rows}</tbody>
  </table>

  <p>Smartsheet-driven tickets account for {sm_total:,} process-required tickets ({sm_pct_pr}). These are going away with Smartsheet retirement, which will independently reduce the process-required category significantly.</p>

  <hr>

  <h2 class="section">Self-Service Detail: The Full Breakdown</h2>

  <table>
    <thead><tr><th>Sub-Category</th><th>Prior</th><th>Current</th><th>Total</th><th>% of SS</th><th>Tool Available</th></tr></thead>
    <tbody>{ss_rows}</tbody>
  </table>

  <hr>

  <h2 class="section">The Bottom Line</h2>

  <p>The week-to-week variance in our ticket volumes is not random and it is not a sign of operational instability. It is the predictable result of Chewtopian self-service behavior &mdash; specifically, how many TMs choose to call HR instead of using available tools on a given week. For now, we are sizing that opportunity at roughly {assumed_ss_pct_of_total:.0f}% of volume, while the rules explicitly confirm {ss_pct_of_total:.1f}% from ticket text alone. The tools exist. The opportunity is adoption and better intake clarity.</p>

  <p>The four focus areas &mdash; timeclock compliance, UKG app and desktop adoption for absences and time-off, balance visibility, and targeted site campaigns &mdash; are actionable, measurable, and tied directly to the data. If we execute on them, we reduce baseline volume, narrow the WoW variance, and shift the OBR conversation from <em>&ldquo;what happened this week?&rdquo;</em> to <em>&ldquo;how much closer are we to our self-service targets?&rdquo;</em></p>

  <hr>

  <p style="font-size:13px;color:var(--muted);line-height:1.5;">
    <strong>Methodology:</strong> {total:,} tickets classified across 4 HR services. Classification is rule-based on ticket description and contact type. Prior Week: {prior_lbl}. Current Week: {curr_lbl}. This version of the report uses a temporary working assumption that 80% of total volume could be self-service, while the rule-based classifier explicitly confirms {ss_pct_of_total:.1f}% from available ticket text. Self-service deflection estimates use conservative 25&ndash;30% adoption assumptions. Report generated automatically by the ORBIT Workload Lens pipeline.
  </p>

  <div class="footer">
    ORBIT HR Workload Lens &nbsp;|&nbsp; Insights &amp; Recommendations &nbsp;|&nbsp; {prior_lbl} to {curr_lbl}<br>
    Generated {report_date} &nbsp;|&nbsp; Pipeline v1.0
  </div>

</div>
</body>
</html>"""

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("  ORBIT Workload Lens — Weekly Pipeline")
    print("=" * 60)

    # 1. Load CSVs
    print(f"\n[1/5] Loading CSVs from: {CSV_DIR}")
    tickets, files_loaded = load_csvs(CSV_DIR)
    for fl in files_loaded:
        print(f"  [OK] {fl}")
    print(f"  Total unique tickets: {len(tickets):,}")

    # 2. Detect weeks
    print(f"\n[2/5] Detecting week boundaries...")
    prior_start, current_start = detect_weeks(tickets)
    prior_end = prior_start + timedelta(days=6)
    current_end = current_start + timedelta(days=6)
    print(f"  Prior Week:   {prior_start.strftime('%b %d')} – {prior_end.strftime('%b %d, %Y')}")
    print(f"  Current Week: {current_start.strftime('%b %d')} – {current_end.strftime('%b %d, %Y')}")

    # Assign weeks
    assign_weeks(tickets, prior_start, current_start)
    in_scope = [t for t in tickets if t.get('week') in ('prior', 'current')]
    print(f"  Tickets in scope (2 weeks): {len(in_scope):,}")

    # 3. Classify
    print(f"\n[3/5] Classifying tickets...")
    classify_all(in_scope)
    class_counts = Counter(t['classification'] for t in in_scope)
    for c, n in class_counts.most_common():
        print(f"  {c}: {n:,} ({n/len(in_scope)*100:.1f}%)")

    # 4. Extract OBR PDF
    print(f"\n[4/5] Checking for OBR PDF...")
    obr_text = extract_obr_text(CSV_DIR)
    if obr_text:
        print(f"  [OK] Extracted {len(obr_text):,} characters from PDF")
    else:
        print(f"  No PDF found (optional — report will generate without it)")

    # 5. Generate report
    print(f"\n[5/5] Generating report...")
    metrics = compute_metrics(in_scope, prior_start, current_start)

    wk_label = current_start.strftime('%Y-%m-%d')
    output_path = os.path.join(OUTPUT_DIR, f"workload_lens_insights_{wk_label}.html")
    generate_report(metrics, obr_text, output_path)
    print(f"  [OK] Report saved to: {output_path}")

    # Also export classified data as CSV
    csv_out = os.path.join(OUTPUT_DIR, f"classified_tickets_{wk_label}.csv")
    with open(csv_out, 'w', encoding='utf-8', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=[
            'number', 'service', 'opened', 'assignment_group', 'contact_type',
            'classification', 'sub_category', 'ss_channel', 'week', 'description'
        ])
        writer.writeheader()
        for t in in_scope:
            writer.writerow({
                'number': t['number'],
                'service': t['service'],
                'opened': t['opened'].strftime('%Y-%m-%d') if t['opened'] else '',
                'assignment_group': t['assignment_group'],
                'contact_type': t['contact_type'],
                'classification': t['classification'],
                'sub_category': t['sub_category'],
                'ss_channel': t['ss_channel'],
                'week': t['week'],
                'description': t['description'][:500],
            })
    print(f"  [OK] Classified data: {csv_out}")

    # Export Self Service tickets to Excel for training review
    ss_xlsx = os.path.join(OUTPUT_DIR, f"self_service_review_{wk_label}.xlsx")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Self Service Review"
        headers = ['Ticket #', 'HR Service', 'Data Source', 'Opened', 'Assignment Group',
                   'Contact Type', 'Sub-Category', 'SS Channel', 'Week', 'Description']
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='0D5F73', end_color='0D5F73', fill_type='solid')
        thin_border = Border(
            bottom=Side(style='thin', color='E2E6EE')
        )
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row_num = 2
        for t in in_scope:
            if t['classification'] != 'Self Service Eligible':
                continue
            ws.cell(row=row_num, column=1, value=t['number'])
            ws.cell(row=row_num, column=2, value=t['service'])
            ws.cell(row=row_num, column=3, value=f"{t['service']} ({t.get('contact_type', '')})")  # data source clarity
            ws.cell(row=row_num, column=4, value=t['opened'].strftime('%Y-%m-%d') if t['opened'] else '')
            ws.cell(row=row_num, column=5, value=t['assignment_group'])
            ws.cell(row=row_num, column=6, value=t['contact_type'])
            ws.cell(row=row_num, column=7, value=t['sub_category'])
            ws.cell(row=row_num, column=8, value=t['ss_channel'])
            ws.cell(row=row_num, column=9, value=t['week'])
            ws.cell(row=row_num, column=10, value=t['description'][:500])
            for col in range(1, 11):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
        # Auto-fit column widths
        col_widths = [14, 24, 32, 12, 28, 20, 40, 20, 10, 80]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        # Freeze header row
        ws.freeze_panes = 'A2'
        # Add auto-filter
        ws.auto_filter.ref = f'A1:J{row_num-1}'
        try:
            wb.save(ss_xlsx)
            print(f"  [OK] Self Service review Excel: {ss_xlsx}")
        except PermissionError:
            alt_ss_xlsx = os.path.join(
                OUTPUT_DIR,
                f"self_service_review_{wk_label}_{datetime.now().strftime('%H%M%S')}.xlsx"
            )
            wb.save(alt_ss_xlsx)
            print(f"  Warning: {ss_xlsx} is open or locked; saved review Excel to: {alt_ss_xlsx}")
    except ImportError:
        print(f"  Warning: openpyxl not installed — skipping Excel export. Install with: pip install openpyxl")

    print(f"\n{'=' * 60}")
    print(f"  Done. Open the HTML report in your browser.")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
