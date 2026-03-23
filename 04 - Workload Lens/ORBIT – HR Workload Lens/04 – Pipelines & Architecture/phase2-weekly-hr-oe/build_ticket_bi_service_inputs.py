from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl


DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%Y-%m-%d",
    "%Y-%m-%d %H:%M:%S",
)

SERVICE_SPECS = [
    {
        "service_name": "Attendance Inquiry",
        "arg_name": "attendance_csv",
        "tokens": ["attendance", "inquiry"],
        "aliases": ["attendance inquiry"],
    },
    {
        "service_name": "CS Time and Attendance",
        "arg_name": "cs_time_attendance_csv",
        "tokens": ["cs", "time", "attendance"],
        "aliases": ["cs time and attendance", "cc time and attendance", "contact center time and attendance"],
    },
    {
        "service_name": "FC General Inquiry",
        "arg_name": "fc_general_inquiry_csv",
        "tokens": ["fc", "general", "inquiry"],
        "aliases": ["fc general inquiry"],
    },
    {
        "service_name": "Timesheet Inquiry",
        "arg_name": "timesheet_inquiry_csv",
        "tokens": ["timesheet", "inquiry"],
        "aliases": ["timesheet inquiry"],
    },
]

OPENED_FIELD_ALIASES = ["Opened At", "Opened", "Created At"]
RESOLVED_FIELD_ALIASES = ["U Resolved", "Resolved At", "Closed At", "Closed", "Resolved"]
HEADER_FAMILIES = {
    "hr_service": ["Hr Service", "Service", "Ticket Type"],
    "ticket_number": ["Number", "Ticket Number", "Case Number"],
    "opened_at": OPENED_FIELD_ALIASES,
    "resolved_at": RESOLVED_FIELD_ALIASES,
    "assignment_group": ["Assignment Group", "AssignmentGroup"],
    "description": ["Description1", "Description", "Short Description"],
}
SOURCE_EXTENSIONS = {".csv", ".xlsx", ".xlsm"}
OPENED_REPORT_TOKENS = [
    "wbr previous week open cases",
    "open cases",
    "cases opened",
    "opened last week",
    "opened",
    "open last week",
]
CLOSED_REPORT_TOKENS = [
    "wbr previous week resolved cases",
    "resolved cases",
    "cases closed",
    "closed last week",
    "closed",
    "resolved",
    "resolve",
]


@dataclass
class ReportMetadata:
    path: Path
    report_type: str
    row_count: int
    valid_metric_rows: int
    metric_start: dt.date | None
    metric_end: dt.date | None
    sheets: list[str]


@dataclass
class SelectedReport:
    metadata: ReportMetadata
    week_label: str
    week_start: dt.date
    week_end: dt.date


def normalize_name(value: str) -> str:
    value = (value or "").lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_header(header: str) -> str:
    header = (header or "").replace("\ufeff", "").strip().lower().replace("_", " ")
    return re.sub(r"\s+", " ", header)


def to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.replace(tzinfo=None).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def parse_datetime(value: object) -> dt.datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, dt.date):
        return dt.datetime.combine(value, dt.time())

    text = to_text(value)
    if not text:
        return None
    text = text.replace("\u200f", "").replace("\u200e", "")
    for fmt in DATE_FORMATS:
        try:
            return dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    if " " in text:
        token = text.split()[0]
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                return dt.datetime.strptime(token, fmt)
            except ValueError:
                pass
    return None


def pick_first(row: dict[str, str], aliases: list[str]) -> str:
    for alias in aliases:
        key = normalize_header(alias)
        if key in row and (row[key] or "").strip():
            return row[key].strip()
    return ""


def looks_like_header(values: list[str]) -> bool:
    normalized = {normalize_header(value) for value in values if value}
    if not normalized:
        return False
    has_ticket = any(normalize_header(alias) in normalized for alias in HEADER_FAMILIES["ticket_number"])
    has_opened = any(normalize_header(alias) in normalized for alias in HEADER_FAMILIES["opened_at"])
    has_service = any(normalize_header(alias) in normalized for alias in HEADER_FAMILIES["hr_service"])
    return has_ticket and (has_opened or has_service)


def read_csv_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for raw_row in reader:
            row = {}
            for key, value in raw_row.items():
                if key is None:
                    continue
                row[to_text(key)] = to_text(value)
            rows.append(row)
    return rows, ["csv"]


def read_workbook_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    used_sheets: list[str] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        headers: list[str] | None = None
        sheet_used = False
        for raw_row in worksheet.iter_rows(values_only=True):
            values = [to_text(value) for value in raw_row]
            if headers is None:
                if looks_like_header(values):
                    headers = values
                    sheet_used = True
                continue
            if not any(values):
                continue
            row = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                row[header] = values[index] if index < len(values) else ""
            rows.append(row)
        if sheet_used:
            used_sheets.append(sheet_name)
    if not rows:
        raise ValueError(f"No tabular ticket rows found in workbook: {path}")
    return rows, used_sheets


def read_tabular_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_workbook_rows(path)
    raise ValueError(f"Unsupported BI report extension: {path.suffix}")


def detect_report_type(path: Path) -> str | None:
    normalized = normalize_name(path.stem)
    if any(token in normalized for token in CLOSED_REPORT_TOKENS):
        return "closed"
    if any(token in normalized for token in OPENED_REPORT_TOKENS):
        return "opened"
    return None


def summarize_report(path: Path, report_type: str) -> ReportMetadata:
    rows, sheets = read_tabular_rows(path)
    metric_aliases = OPENED_FIELD_ALIASES if report_type == "opened" else RESOLVED_FIELD_ALIASES
    metric_dates: list[dt.date] = []
    for raw_row in rows:
        normalized = {normalize_header(key): to_text(value) for key, value in raw_row.items() if key is not None}
        parsed = parse_datetime(pick_first(normalized, metric_aliases))
        if parsed:
            metric_dates.append(parsed.date())

    metric_dates.sort()
    return ReportMetadata(
        path=path.resolve(),
        report_type=report_type,
        row_count=len(rows),
        valid_metric_rows=len(metric_dates),
        metric_start=metric_dates[0] if metric_dates else None,
        metric_end=metric_dates[-1] if metric_dates else None,
        sheets=sheets,
    )


def discover_reports(bi_weekly_dir: Path) -> list[ReportMetadata]:
    reports: list[ReportMetadata] = []
    for path in sorted(bi_weekly_dir.iterdir()):
        if not path.is_file() or path.suffix.lower() not in SOURCE_EXTENSIONS:
            continue
        report_type = detect_report_type(path)
        if not report_type:
            continue
        reports.append(summarize_report(path, report_type))
    return reports


def latest_completed_saturday(pull_date: dt.date) -> dt.date:
    days_since_saturday = (pull_date.weekday() - 5) % 7
    return pull_date - dt.timedelta(days=days_since_saturday)


def score_report_match(report: ReportMetadata, week_start: dt.date, week_end: dt.date) -> tuple[int, int, float]:
    if report.metric_start is None or report.metric_end is None:
        return (0, 0, report.path.stat().st_mtime)
    if report.metric_start == week_start and report.metric_end == week_end:
        return (3, 0, report.path.stat().st_mtime)
    if report.metric_start <= week_start and report.metric_end >= week_end:
        span_days = (report.metric_end - report.metric_start).days
        return (2, -span_days, report.path.stat().st_mtime)
    if week_start <= report.metric_start <= week_end or week_start <= report.metric_end <= week_end:
        overlap_days = min(report.metric_end, week_end) - max(report.metric_start, week_start)
        return (1, overlap_days.days if isinstance(overlap_days, dt.timedelta) else 0, report.path.stat().st_mtime)
    return (0, 0, report.path.stat().st_mtime)


def select_report_for_week(
    reports: list[ReportMetadata],
    report_type: str,
    week_label: str,
    week_start: dt.date,
    week_end: dt.date,
) -> SelectedReport:
    candidates = [report for report in reports if report.report_type == report_type]
    ranked = sorted(
        candidates,
        key=lambda report: score_report_match(report, week_start, week_end),
        reverse=True,
    )
    if not ranked or score_report_match(ranked[0], week_start, week_end)[0] == 0:
        raise FileNotFoundError(
            f"Could not find a {report_type} BI report covering {week_label} "
            f"({week_start.isoformat()} to {week_end.isoformat()}) in the supplied folder."
        )
    return SelectedReport(
        metadata=ranked[0],
        week_label=week_label,
        week_start=week_start,
        week_end=week_end,
    )


def normalize_service_name(raw_value: str) -> str | None:
    normalized = normalize_name(raw_value)
    if not normalized:
        return None
    for spec in SERVICE_SPECS:
        if normalized in spec["aliases"]:
            return spec["service_name"]
        if all(token in normalized for token in spec["tokens"]):
            return spec["service_name"]
    return None


def choose_text_value(current: str, new_value: str) -> str:
    if not current:
        return new_value
    if not new_value:
        return current
    return new_value if len(new_value) > len(current) else current


def merge_record_value(existing: dict[str, object], field: str, new_value: object) -> None:
    current = existing.get(field)
    if field == "opened_at":
        if new_value and (current is None or new_value < current):
            existing[field] = new_value
        return
    if field == "resolved_at":
        if new_value and (current is None or new_value > current):
            existing[field] = new_value
        return
    if field in {"assignment_group", "description"}:
        existing[field] = choose_text_value(str(current or ""), str(new_value or ""))
        return
    if field == "ticket_number":
        if not current and new_value:
            existing[field] = new_value
        return
    if field == "service_name":
        if not current and new_value:
            existing[field] = new_value


def format_datetime(value: dt.datetime | None) -> str:
    if not value:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S")


def to_slug(text: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return slug or "dataset"


def build_service_inputs(
    selected_reports: list[SelectedReport],
    out_dir: Path,
) -> dict[str, object]:
    out_dir.mkdir(parents=True, exist_ok=True)
    records_by_key: dict[tuple[str, str], dict[str, object]] = {}
    report_processing: list[dict[str, object]] = []
    skipped_unknown_service = 0

    for selected in selected_reports:
        rows, sheets = read_tabular_rows(selected.metadata.path)
        rows_in_window = 0
        rows_in_scope = 0
        for index, raw_row in enumerate(rows, start=1):
            normalized = {normalize_header(key): to_text(value) for key, value in raw_row.items() if key is not None}
            opened_at = parse_datetime(pick_first(normalized, OPENED_FIELD_ALIASES))
            resolved_at = parse_datetime(pick_first(normalized, RESOLVED_FIELD_ALIASES))
            metric_dt = opened_at if selected.metadata.report_type == "opened" else resolved_at
            if metric_dt is None:
                continue
            metric_date = metric_dt.date()
            if not (selected.week_start <= metric_date <= selected.week_end):
                continue
            rows_in_window += 1

            service_name = normalize_service_name(pick_first(normalized, HEADER_FAMILIES["hr_service"]))
            if service_name is None:
                skipped_unknown_service += 1
                continue

            ticket_number = pick_first(normalized, HEADER_FAMILIES["ticket_number"]) or (
                f"{selected.metadata.path.stem}:{selected.week_label}:{index}"
            )
            record_key = (service_name, ticket_number)
            record = records_by_key.setdefault(
                record_key,
                {
                    "service_name": service_name,
                    "ticket_number": ticket_number,
                    "opened_at": None,
                    "resolved_at": None,
                    "assignment_group": "",
                    "description": "",
                    "source_report_types": set(),
                    "source_report_weeks": set(),
                    "source_files": set(),
                },
            )
            merge_record_value(record, "service_name", service_name)
            merge_record_value(record, "ticket_number", ticket_number)
            merge_record_value(record, "opened_at", opened_at)
            merge_record_value(record, "resolved_at", resolved_at)
            merge_record_value(record, "assignment_group", pick_first(normalized, HEADER_FAMILIES["assignment_group"]))
            merge_record_value(record, "description", pick_first(normalized, HEADER_FAMILIES["description"]))
            record["source_report_types"].add(selected.metadata.report_type)
            record["source_report_weeks"].add(f"{selected.week_start.isoformat()}_to_{selected.week_end.isoformat()}")
            record["source_files"].add(str(selected.metadata.path))
            rows_in_scope += 1

        report_processing.append(
            {
                "path": str(selected.metadata.path),
                "report_type": selected.metadata.report_type,
                "week_label": selected.week_label,
                "week_start": selected.week_start.isoformat(),
                "week_end": selected.week_end.isoformat(),
                "sheets": sheets,
                "rows_in_window": rows_in_window,
                "rows_in_scope": rows_in_scope,
            }
        )

    written_service_files: list[dict[str, object]] = []
    for spec in SERVICE_SPECS:
        service_rows = [
            record
            for record in records_by_key.values()
            if record["service_name"] == spec["service_name"]
        ]
        service_rows.sort(
            key=lambda item: (
                item["opened_at"] or dt.datetime.min,
                str(item["ticket_number"]),
            )
        )
        service_path = out_dir / f"{to_slug(spec['service_name'])}_from_bi_weekly_reports.csv"
        with service_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "Hr Service",
                    "Opened At",
                    "Number",
                    "Assignment Group",
                    "Description1",
                    "U Resolved",
                    "Source Report Types",
                    "Source Report Weeks",
                    "Source Files",
                ],
            )
            writer.writeheader()
            for row in service_rows:
                writer.writerow(
                    {
                        "Hr Service": row["service_name"],
                        "Opened At": format_datetime(row["opened_at"]),
                        "Number": row["ticket_number"],
                        "Assignment Group": row["assignment_group"],
                        "Description1": row["description"],
                        "U Resolved": format_datetime(row["resolved_at"]),
                        "Source Report Types": ", ".join(sorted(row["source_report_types"])),
                        "Source Report Weeks": ", ".join(sorted(row["source_report_weeks"])),
                        "Source Files": " | ".join(sorted(row["source_files"])),
                    }
                )
        written_service_files.append(
            {
                "service_name": spec["service_name"],
                "arg_name": spec["arg_name"],
                "path": str(service_path),
                "row_count": len(service_rows),
            }
        )

    manifest = {
        "pass": True,
        "source_mode": "bi_weekly_two_report_folder",
        "report_processing": report_processing,
        "skipped_unknown_service_rows": skipped_unknown_service,
        "service_csvs": written_service_files,
    }
    manifest_path = out_dir / "bi_weekly_ticket_intake_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    return {
        **manifest,
        "manifest_json": str(manifest_path),
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build canonical per-service ticket CSVs from the weekly BI opened/closed report folder."
    )
    parser.add_argument("--bi-weekly-dir", required=True)
    parser.add_argument("--pull-date", default=dt.date.today().isoformat())
    parser.add_argument("--week8-start")
    parser.add_argument("--week8-end")
    parser.add_argument("--week9-start")
    parser.add_argument("--week9-end")
    parser.add_argument("--out-dir", required=True)
    args = parser.parse_args()

    pull_date = dt.date.fromisoformat(args.pull_date)
    required_week9_end = latest_completed_saturday(pull_date)
    required_week9_start = required_week9_end - dt.timedelta(days=6)
    required_week8_end = required_week9_start - dt.timedelta(days=1)
    required_week8_start = required_week8_end - dt.timedelta(days=6)

    week8_start = dt.date.fromisoformat(args.week8_start) if args.week8_start else required_week8_start
    week8_end = dt.date.fromisoformat(args.week8_end) if args.week8_end else required_week8_end
    week9_start = dt.date.fromisoformat(args.week9_start) if args.week9_start else required_week9_start
    week9_end = dt.date.fromisoformat(args.week9_end) if args.week9_end else required_week9_end

    bi_weekly_dir = Path(args.bi_weekly_dir).resolve()
    if not bi_weekly_dir.exists():
        raise SystemExit(f"BI weekly folder not found: {bi_weekly_dir}")

    reports = discover_reports(bi_weekly_dir)
    if not reports:
        raise SystemExit(
            "No BI weekly ticket files were discovered. Expected CSV/XLSX files with open/opened or resolved/closed cues in the name."
        )

    selected_reports = [
        select_report_for_week(reports, "opened", "week8", week8_start, week8_end),
        select_report_for_week(reports, "closed", "week8", week8_start, week8_end),
        select_report_for_week(reports, "opened", "week9", week9_start, week9_end),
        select_report_for_week(reports, "closed", "week9", week9_start, week9_end),
    ]

    out_dir = Path(args.out_dir).resolve()
    payload = build_service_inputs(selected_reports, out_dir)
    payload["selected_reports"] = [
        {
            "path": str(selected.metadata.path),
            "report_type": selected.metadata.report_type,
            "week_label": selected.week_label,
            "week_start": selected.week_start.isoformat(),
            "week_end": selected.week_end.isoformat(),
            "metric_start": selected.metadata.metric_start.isoformat() if selected.metadata.metric_start else None,
            "metric_end": selected.metadata.metric_end.isoformat() if selected.metadata.metric_end else None,
            "row_count": selected.metadata.row_count,
            "valid_metric_rows": selected.metadata.valid_metric_rows,
        }
        for selected in selected_reports
    ]
    payload["week_lock"] = {
        "week8_start": week8_start.isoformat(),
        "week8_end": week8_end.isoformat(),
        "week9_start": week9_start.isoformat(),
        "week9_end": week9_end.isoformat(),
        "definition": "Sunday through Saturday",
    }
    payload["pull_date"] = pull_date.isoformat()
    manifest_path = Path(payload["manifest_json"])
    manifest_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps(payload))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
